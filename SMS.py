import tkinter as tk
from tkinter import ttk, messagebox
import serial
import serial.tools.list_ports
import re
import subprocess
import time
import concurrent.futures
import threading
import json
import logging
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk, ImageDraw, ImageFont
from firmware_version_integration import check_firmware_version
import tkinter.simpledialog as simpledialog

# Get current date for folder naming and filename
current_date = datetime.now().strftime("%Y-%m-%d")

# Setup directories
Test_Result = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Test Result Excel File")
Log_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Log files")

os.makedirs(Test_Result, exist_ok=True)
os.makedirs(Log_folder, exist_ok=True)

log_folder = os.path.join(Log_folder, f"Log_file_{current_date}")
excel_folder = os.path.join(Test_Result, f"Test_Result_{current_date}")

os.makedirs(log_folder, exist_ok=True)
os.makedirs(excel_folder, exist_ok=True)

file_name = os.path.join(excel_folder, f"{current_date}_NIC_RF_TEST_RESULT.xlsx")

if os.path.exists(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIC serial number", "RF test result", "RX RSSI", "TX RSSI", "Action", "Timestamp"])

def log_action(action, timestamp, serial_number=None):
    log_file_path = os.path.join(log_folder, "action_log.txt")
    with open(log_file_path, "a") as log_file:
        if serial_number:
            log_file.write(f"{timestamp} - Serial Number: {serial_number} - {action}\n")
        else:
            log_file.write(f"{timestamp} - {action}\n")

logging.basicConfig(
    filename=os.path.join(log_folder, f"{current_date}_program_log.txt"),
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

class SerialScanner:
    """Enhanced serial scanning functionality with retry mechanism"""
    
    def __init__(self, port_config, baud_rate=9600, trigger_command=b'*T', timeout=3, max_retries=1):
        self.port_config = port_config
        self.baud_rate = baud_rate
        self.trigger_command = trigger_command
        self.timeout = timeout
        self.max_retries = max_retries
        self.scan_results = {}
        self.scan_lock = threading.Lock()
        self.retry_stations = set()  # Track stations that need retry
    
    def scan_single_port(self, station_num, port, is_retry=False):
        """Scan a single COM port for serial number"""
        retry_text = " (RETRY)" if is_retry else ""
        try:
            logging.info(f"[Station {station_num}] Scanning port {port}{retry_text}")
            with serial.Serial(port, self.baud_rate, timeout=self.timeout) as ser:
                # Send trigger command
                ser.write(self.trigger_command)
                logging.debug(f"[Station {station_num}] Sent trigger command to {port}{retry_text}")
                
                # Wait for response
                time.sleep(1)
                scanned_data = ser.readline().decode('utf-8', errors='ignore').strip()
                
                with self.scan_lock:
                    if scanned_data:
                        self.scan_results[station_num] = {
                            'serial': scanned_data,
                            'port': port,
                            'status': 'success',
                            'retry_attempt': is_retry
                        }
                        # Remove from retry list if successful
                        self.retry_stations.discard(station_num)
                        logging.info(f"[Station {station_num}] Successfully scanned{retry_text}: {scanned_data}")
                    else:
                        self.scan_results[station_num] = {
                            'serial': None,
                            'port': port,
                            'status': 'no_data',
                            'retry_attempt': is_retry
                        }
                        # Add to retry list if first attempt failed
                        if not is_retry:
                            self.retry_stations.add(station_num)
                        logging.warning(f"[Station {station_num}] No data received from {port}{retry_text}")
                        
        except serial.SerialException as e:
            with self.scan_lock:
                self.scan_results[station_num] = {
                    'serial': None,
                    'port': port,
                    'status': 'error',
                    'error': str(e),
                    'retry_attempt': is_retry
                }
                # Add to retry list if first attempt failed
                if not is_retry:
                    self.retry_stations.add(station_num)
            logging.error(f"[Station {station_num}] Serial error on {port}{retry_text}: {e}")
        except Exception as e:
            with self.scan_lock:
                self.scan_results[station_num] = {
                    'serial': None,
                    'port': port,
                    'status': 'error',
                    'error': str(e),
                    'retry_attempt': is_retry
                }
                # Add to retry list if first attempt failed
                if not is_retry:
                    self.retry_stations.add(station_num)
            logging.error(f"[Station {station_num}] Unexpected error on {port}{retry_text}: {e}")
    
    def scan_all_configured_ports(self):
        """Scan all configured COM ports in parallel with retry mechanism"""
        self.scan_results.clear()
        self.retry_stations.clear()
        
        # First attempt - scan all configured ports
        logging.info("Starting initial scan of all configured ports")
        self._perform_scan_batch(is_retry=False)
        
        # Retry failed stations once
        if self.retry_stations and self.max_retries > 0:
            logging.info(f"Retrying failed stations: {list(self.retry_stations)}")
            time.sleep(2)  # Brief delay before retry
            self._perform_scan_batch(is_retry=True, stations_to_scan=self.retry_stations.copy())
        
        return self.scan_results.copy()
    
    def _perform_scan_batch(self, is_retry=False, stations_to_scan=None):
        """Perform a batch of scans (either initial or retry)"""
        threads = []
        
        # Determine which stations to scan
        if stations_to_scan is None:
            stations_to_scan = self.port_config.keys()
        
        for station_num in stations_to_scan:
            port = self.port_config.get(station_num)
            if port:  # Only scan if port is configured
                thread = threading.Thread(
                    target=self.scan_single_port, 
                    args=(station_num, port, is_retry),
                    daemon=True
                )
                threads.append(thread)
                thread.start()
        
        # Wait for all scans to complete
        for thread in threads:
            thread.join()
        
        return self.scan_results.copy()

class ModernUI:
    """Helper class for modern UI elements"""
    
    @staticmethod
    def create_rounded_rectangle(canvas, x1, y1, x2, y2, radius=25, **kwargs):
        """Draw a rounded rectangle on a canvas"""
        points = [
            x1+radius, y1,
            x2-radius, y1,
            x2, y1,
            x2, y1+radius,
            x2, y2-radius,
            x2, y2,
            x2-radius, y2,
            x1+radius, y2,
            x1, y2,
            x1, y2-radius,
            x1, y1+radius,
            x1, y1
        ]
        return canvas.create_polygon(points, **kwargs, smooth=True)
    
    @staticmethod
    def create_status_indicator(parent, size=60, initial_state="pending"):
        """Create a modern status indicator"""
        canvas = tk.Canvas(parent, width=size, height=size, bg=parent["bg"], highlightthickness=0)
        
        states = {
            "pending": {"fill": "#E0E0E0", "outline": "#BDBDBD", "symbol": "⏱", "symbol_color": "#757575"},
            "scanning": {"fill": "#FFE082", "outline": "#FFC107", "symbol": "📡", "symbol_color": "#F57F17"},
            "retrying": {"fill": "#FFCC80", "outline": "#FF9800", "symbol": "🔄", "symbol_color": "#E65100"},
            "running": {"fill": "#90CAF9", "outline": "#2196F3", "symbol": "⏳", "symbol_color": "#0D47A1"},
            "success": {"fill": "#A5D6A7", "outline": "#4CAF50", "symbol": "✓", "symbol_color": "#1B5E20"},
            "failure": {"fill": "#EF9A9A", "outline": "#F44336", "symbol": "✕", "symbol_color": "#B71C1C"}
        }
        
        # Draw outer circle with shadow effect
        canvas.create_oval(3, 3, size-1, size-1, fill="#F5F5F5", outline="#E0E0E0", width=2)
        
        # Draw inner circle based on state
        state_config = states.get(initial_state, states["pending"])
        inner_size = 4
        canvas.create_oval(inner_size, inner_size, size-inner_size, size-inner_size, 
                          fill=state_config["fill"], outline=state_config["outline"], width=2)
        
        # Add symbol
        canvas.create_text(size/2, size/2, text=state_config["symbol"], 
                          fill=state_config["symbol_color"], font=("Arial", int(size/2), "bold"))
        
        return canvas

    @staticmethod
    def update_status_indicator(canvas, state):
        """Update the status indicator to a new state"""
        size = canvas.winfo_width()
        if size < 10:
            size = 60
        
        # Clear canvas
        canvas.delete("all")
        
        states = {
            "pending": {"fill": "#E0E0E0", "outline": "#BDBDBD", "symbol": "⏱", "symbol_color": "#757575"},
            "scanning": {"fill": "#FFE082", "outline": "#FFC107", "symbol": "📡", "symbol_color": "#F57F17"},
            "retrying": {"fill": "#FFCC80", "outline": "#FF9800", "symbol": "🔄", "symbol_color": "#E65100"},
            "running": {"fill": "#90CAF9", "outline": "#2196F3", "symbol": "⏳", "symbol_color": "#0D47A1"},
            "success": {"fill": "#4CAF50", "outline": "#2E7D32", "symbol": "✓", "symbol_color": "#FFFFFF"},
            "failure": {"fill": "#F44336", "outline": "#B71C1C", "symbol": "✕", "symbol_color": "#FFFFFF"}
        }
        
        state_config = states.get(state, states["pending"])
        
        # Draw outer circle with gradient effect
        gradient_colors = ["#D0D0D0", "#A0A0A0"]
        steps = 10
        step_size = (size - 4) / (2 * steps)
        
        for i in range(steps):
            color = gradient_colors[i % 2]
            canvas.create_oval(2 + i * step_size, 2 + i * step_size, size - 2 - i * step_size, size - 2 - i * step_size,
                            outline=color, width=1)
        
        # Draw inner circle with shadow effect
        gradient_offset = 4
        canvas.create_oval(gradient_offset, gradient_offset, size - gradient_offset, size - gradient_offset,
                        fill=state_config["fill"], outline=state_config["outline"], width=2)
        
        # Add glossy highlight effect
        highlight_offset = 8
        canvas.create_arc(highlight_offset, highlight_offset, size - highlight_offset, size - highlight_offset,
                        start=30, extent=120, style="arc", outline="#FFFFFF", width=2)
        
        # Add subtle shadow for 3D effect
        shadow_offset = 1
        canvas.create_oval(gradient_offset + shadow_offset, gradient_offset + shadow_offset,
                        size - gradient_offset - shadow_offset, size - gradient_offset - shadow_offset,
                        outline="#808080", width=1)
        
        # Add symbol
        canvas.create_text(size / 2, size / 2, text=state_config["symbol"],
                        fill=state_config["symbol_color"], font=("Arial", int(size / 2.5), "bold"))

    @staticmethod
    def create_modern_button(parent, text, command, width=120, height=40, bg="#4CAF50", hover_bg="#388E3C"):
        def on_enter(e):
            button["bg"] = hover_bg
            button["relief"] = "raised"
        
        def on_leave(e):
            button["bg"] = bg
            button["relief"] = "ridge"
        
        def on_press(e):
            button["relief"] = "sunken"
        
        def on_release(e):
            button["relief"] = "raised"

        button = tk.Label(parent, text=text, width=width, height=height,
                          bg=bg, fg="white", font=("Arial", 12, "bold"),
                          bd=3, relief="ridge", padx=5, pady=5, cursor="hand2")

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
        button.bind("<ButtonPress-1>", on_press)
        button.bind("<ButtonRelease-1>", lambda e: (on_release(e), command()))

        button.pack_propagate(False)
        return button
    
    @staticmethod
    def create_task_indicator(parent, text, initial_state="pending"):
        """Create a task indicator with icon and text"""
        frame = tk.Frame(parent, bg=parent["bg"])
        
        states = {
            "pending": {"symbol": "⏱", "color": "#757575"},
            "running": {"symbol": "⏳", "color": "#2196F3"},
            "success": {"symbol": "✓", "color": "#4CAF50"},
            "failure": {"symbol": "✕", "color": "#F44336"}
        }
        
        state_config = states.get(initial_state, states["pending"])
        
        # Create icon label
        icon_label = tk.Label(frame, text=state_config["symbol"], fg=state_config["color"], 
                             bg=parent["bg"], font=("Arial", 12, "bold"))
        icon_label.pack(side=tk.LEFT, padx=(0, 5))
        
        # Create text label
        text_label = tk.Label(frame, text=text, anchor="w", bg=parent["bg"], font=("Arial", 11))
        text_label.pack(side=tk.LEFT, fill=tk.X)
        
        return frame, icon_label

class ManufacturingSuite:
    def __init__(self, root):
        # Initialize the main window
        self.root = root
        self.root.title("Smart Manufacturing Suite - [SMS] [Version 1.0.0.0, 06-March-25]")
        self.root.geometry("1350x700")
        self.root.configure(bg="#F5F5F5")
        
        # Set to full screen
        self.root.attributes('-fullscreen', True)
        self.root.bind("<Escape>", self.exit_fullscreen)
        
        # Load configuration for COM ports mapping
        config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.txt")
        self.port_config = self.load_port_config(config_path)
        
        # Initialize serial scanner with retry capability
        self.serial_scanner = SerialScanner(self.port_config, max_retries=1)
        
        self.active_ports_serials = []
        self.task_labels = {}
        self.station_numbers = {}
        self.entries = {}
        self.scanned_serials = set()
        self.eng_var = tk.BooleanVar()
        self.status_indicators = {}
        self.task_indicators = {}
        
        # NEW: Track failed stations
        self.failed_stations = set()
        
        # NEW: Add synchronization for test completion tracking
        self.test_completion_lock = threading.Lock()
        self.completed_tests = {}  # Track completion status of each test type
        self.station_test_results = {}  # Store individual test results per station
        
        # Create the UI components
        self.create_header()
        self.create_main_content()
        self.create_status_bar()
        
        # Timer variables
        self.start_time = None
        self.timer_running = False
        self.timer_id = None
        
        # Load channel data
        self.load_channel_data()

    def load_port_config(self, path):
        """Load COM port configuration from config.txt"""
        config = {}
        if os.path.exists(path):
            try:
                with open(path, "r") as f:
                    for line in f:
                        line = line.strip()
                        if "=" in line and not line.startswith("#"):
                            station, port = line.split("=", 1)
                            config[int(station.strip())] = port.strip()
                logging.info(f"Loaded port configuration: {config}")
            except Exception as e:
                logging.error(f"Error loading port configuration: {e}")
                messagebox.showerror("Config Error", f"Failed to load port configuration: {e}")
        else:
            logging.warning(f"COM port config file not found at: {path}")
            messagebox.showwarning("Config Missing", 
                                 f"Configuration file not found: {path}\n"
                                 "Please create config.txt with station=COMport mappings")
        return config

    def auto_scan_all_serials(self):
        """Auto scan serial numbers from all configured COM ports with enhanced retry mechanism"""
        self.status_label.config(text="Auto-scanning serial numbers...")
        
        # Update status indicators to show scanning
        for station_num in self.station_numbers.keys():
            if station_num in self.port_config:
                ModernUI.update_status_indicator(self.status_indicators[station_num], "scanning")
        
        self.root.update()
        
        # Perform the scan
        scan_results = self.serial_scanner.scan_all_configured_ports()
        
        # Process scan results
        successful_scans = 0
        failed_scans = 0
        retried_scans = 0
        
        for station_num, result in scan_results.items():
            if result['status'] == 'success' and result['serial']:
                # Update the entry field with scanned serial
                self.station_numbers[station_num].set(result['serial'])
                self.finalize_serial_entry(station_num)
                successful_scans += 1
                
                # Log if this was a retry success
                if result.get('retry_attempt', False):
                    retried_scans += 1
                    logging.info(f"Station {station_num}: Retry successful - Scanned {result['serial']} from {result['port']}")
                else:
                    logging.info(f"Station {station_num}: Scanned {result['serial']} from {result['port']}")
            else:
                # Handle scan failure (after retry attempts)
                failed_scans += 1
                self.failed_stations.add(station_num)
                error_msg = result.get('error', 'No data received')
                retry_text = " (after retry)" if result.get('retry_attempt', False) else ""
                
                # Log the failure
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_action(f"Station {station_num} scan failed{retry_text}: {error_msg}", timestamp)
                
                # Add failed scan to Excel
                sheet.append([
                    f"Station_{station_num}_SCAN_FAILED", 
                    "Fail", 
                    "N/A", 
                    "N/A", 
                    f"Serial scan failed{retry_text}: {error_msg}", 
                    timestamp
                ])
                workbook.save(file_name)
                
                # Update UI to show failure
                logging.warning(f"Station {station_num}: Scan failed{retry_text} - {error_msg}")
                ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                
                # Mark all tasks as failed for this station
                for task_name in self.task_indicators[station_num]:
                    self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
        
        # Update status with detailed information
        status_parts = []
        if successful_scans > 0:
            status_parts.append(f"{successful_scans} successful")
        if retried_scans > 0:
            status_parts.append(f"{retried_scans} recovered on retry")
        if failed_scans > 0:
            status_parts.append(f"{failed_scans} failed")
        
        if status_parts:
            status_msg = f"Auto-scan completed: {', '.join(status_parts)}"
        else:
            status_msg = "Auto-scan completed: No serial numbers found"
        
        self.status_label.config(text=status_msg)
            
        # Log comprehensive summary
        logging.info(f"Scan summary: {successful_scans} successful ({retried_scans} after retry), {failed_scans} failed")
        if failed_scans > 0:
            logging.warning(f"Failed stations (after retry): {list(self.failed_stations)}")
        if retried_scans > 0:
            logging.info(f"Stations recovered on retry: {retried_scans}")

    def finalize_serial_entry(self, station_num):
        """Finalize a serial entry (lock the field and update status)"""
        entry_widget = self.entries[station_num]
        serial_value = self.station_numbers[station_num].get().strip()
        
        if serial_value:
            entry_widget.config(state="disabled", disabledbackground="#F5F5F5", disabledforeground="#333")
            self.scanned_serials.add(serial_value)
            ModernUI.update_status_indicator(self.status_indicators[station_num], "pending")

    def handle_manual_entry(self, event, station_num):
        """Handle manual serial number entry"""
        serial_value = self.station_numbers[station_num].get().strip()
        if serial_value:
            # Remove from failed stations if manually entered
            self.failed_stations.discard(station_num)
            self.finalize_serial_entry(station_num)

    def exit_fullscreen(self, event=None):
        self.root.attributes('-fullscreen', False)
        
    def load_channel_data(self):
        """Load channel data from configuration file"""
        try:
            with open('Load_channel_data.json', 'r') as file:
                criteria = json.load(file)
                
            self.ENCRYPTION_KEY = criteria.get('ENCRYPTION_KEY', '')
            self.AUTHENTICATION_KEY = criteria.get('AUTHENTICATION_KEY', '')
            self.NETWORK_ADDRESS = criteria.get('NETWORK_ADDRESS', '')
            self.TEST_ROUTER_ADDRESS = criteria.get('TEST_ROUTER_ADDRESS', '')
            self.NETWORK_CHANNEL = criteria.get('NETWORK_CHANNEL', '')
            
            logging.info("Channel data loaded successfully")
        except Exception as e:
            logging.error(f"Error loading channel data: {e}")
            self.ENCRYPTION_KEY = ''
            self.AUTHENTICATION_KEY = ''
            self.NETWORK_ADDRESS = ''
            self.TEST_ROUTER_ADDRESS = ''
            self.NETWORK_CHANNEL = ''
            messagebox.showwarning("Configuration Error", 
                                  "Failed to load channel data. Please configure test router first.")
        
    def create_header(self):
        # Create a frame for the header
        header_frame = tk.Frame(self.root, bg="#1976D2", height=72)
        header_frame.pack(fill=tk.X)
        
        # Logo frame on the left
        logo_frame = tk.Frame(header_frame, bg="#1976D2", width=200)
        logo_frame.pack(side=tk.LEFT, padx=20, pady=10)
        
        # Try to load and display the logo
        if os.path.exists("logo.png"):
            original_image = Image.open("logo.png")
            resized_image = original_image.resize((125, 40))
            logo_image = ImageTk.PhotoImage(resized_image)
            logo_label = tk.Label(logo_frame, image=logo_image, bg="#1976D2")
            logo_label.image = logo_image
        else:
            logo_label = tk.Label(logo_frame, text="POLARIS", font=("Arial", 20, "bold"), 
                                bg="#1976D2", fg="white")

        logo_label.pack(side=tk.LEFT)
        
        # Add title
        title_label = tk.Label(header_frame, text="Smart Manufacturing Suite", 
                              font=("Arial", 18, "bold"), bg="#1976D2", fg="white")
        title_label.pack(side=tk.LEFT, padx=5, pady=15)
        
        # Add version info
        version_label = tk.Label(header_frame, text="v1.0.0.0 | 06-March-25", 
                                font=("Arial", 10), bg="#1976D2", fg="#E1F5FE")
        version_label.pack(side=tk.LEFT, padx=5, pady=15)
        
        # Add engineering mode checkbox
        eng_frame = tk.Frame(header_frame, bg="#1976D2")
        eng_frame.pack(side=tk.RIGHT, padx=20)
        
        eng_check = tk.Checkbutton(eng_frame, text="Engineering Mode", variable=self.eng_var, 
                                  bg="#1976D2", fg="white", selectcolor="#0D47A1", 
                                  activebackground="#1976D2", activeforeground="white")
        eng_check.pack(pady=15)
        
        # Add toolbar
        toolbar_frame = tk.Frame(self.root, bg="#2196F3", height=55)
        toolbar_frame.pack(fill=tk.X)
        
        # Add buttons to toolbar
        run_btn_frame = ModernUI.create_modern_button(toolbar_frame, "▶ Run Test", self.run_test, 
                                                width=10, height=1, bg="#4CAF50", hover_bg="#388E3C")
        run_btn_frame.pack(side=tk.LEFT, padx=15, pady=7)

        
        # Bind Enter key to run_test function
        self.root.bind('<Return>', lambda event: self.run_test())

        stop_btn_frame = ModernUI.create_modern_button(toolbar_frame, "■ Stop", self.stop_test, 
                                                width=10, height=1, bg="#F44336", hover_bg="#D32F2F")
        stop_btn_frame.pack(side=tk.LEFT, padx=15, pady=7)

        new_batch_btn_frame = ModernUI.create_modern_button(toolbar_frame, "New Batch", self.show_start_batch_dialog, 
                                                        width=10, height=1, bg="#555259", hover_bg="#3b393d")
        new_batch_btn_frame.pack(side=tk.LEFT, padx=15, pady=7)
        
    def create_main_content(self):
        # Create a frame for the main content with some padding
        main_frame = tk.Frame(self.root, bg="#F5F5F5")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        # Create a frame for the test stations
        content_frame = tk.Frame(main_frame, bg="#F5F5F5")
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Dictionary to store serial numbers for each test station
        self.station_numbers = {}
        self.task_indicators = {}

        # Create test station frames in a 2x3 grid
        for i in range(2):
            for j in range(3):
                station_num = i * 3 + j + 1
                
                # Create a frame for each test station
                station_frame = tk.Frame(content_frame, bg="white", bd=0)
                station_frame.grid(row=i, column=j, padx=10, pady=10, sticky="nsew")
                
                # Add a canvas for rounded corners and shadow effect
                station_canvas = tk.Canvas(station_frame, bg="white", highlightthickness=0)
                station_canvas.pack(fill="both", expand=True)
                
                # Draw rounded rectangle for the station
                ModernUI.create_rounded_rectangle(station_canvas, 2, 2, 438, 308, radius=10, 
                                                fill="white", outline="#E0E0E0", width=2)
                
                # Create inner frame for content
                inner_frame = tk.Frame(station_canvas, bg="white")
                station_canvas.create_window(200, 145, window=inner_frame, width=380, height=280)
                
                # Station header
                header_frame = tk.Frame(inner_frame, bg="#2196F3", height=40)
                header_frame.pack(fill=tk.X)
                
                station_label = tk.Label(header_frame, text=f"Test Station {station_num}", 
                                        font=("Arial", 12, "bold"), bg="#2196F3", fg="white")
                station_label.pack(side=tk.LEFT, padx=15, pady=8)
                
                # Show configured COM port
                if station_num in self.port_config:
                    port_label = tk.Label(header_frame, text=f"({self.port_config[station_num]})", 
                                        font=("Arial", 10), bg="#2196F3", fg="#E1F5FE")
                    port_label.pack(side=tk.RIGHT, padx=15, pady=8)
                
                # Create status indicator
                status_indicator = ModernUI.create_status_indicator(station_canvas, size=85, initial_state="pending")
                status_indicator.place(x=325, y=120)
                self.status_indicators[station_num] = status_indicator
                
                # Serial number frame
                serial_frame = tk.Frame(inner_frame, bg="white", height=50)
                serial_frame.pack(fill=tk.X, padx=10, pady=(15, 5))
                
                serial_label = tk.Label(serial_frame, text="Serial Number:", bg="white", font=("Arial", 11))
                serial_label.pack(side=tk.LEFT, padx=(5, 10))
                
                # Entry field for serial number
                self.station_numbers[station_num] = tk.StringVar()
                entry = tk.Entry(serial_frame, textvariable=self.station_numbers[station_num], 
                               width=30, font=("Arial", 11), bd=2, relief=tk.GROOVE)
                entry.pack(side=tk.LEFT, padx=5)
                entry.bind("<Return>", lambda event, s=station_num: self.handle_manual_entry(event, s))
                self.entries[station_num] = entry
                
                # Task status frame
                task_frame = tk.Frame(inner_frame, bg="white")
                task_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
                
                # List of tasks
                tasks = [
                    "Mapped with COM Port",
                    "Test Firmware Download",
                    "External Flash Test",
                    "RSSI Test",
                    "Main Firmware Download",
                    "Main Firmware Verification"
                ]
                
                # Store task indicators for this station
                self.task_indicators[station_num] = {}
                
                # Display tasks with status indicators
                for task in tasks:
                    task_indicator, icon_label = ModernUI.create_task_indicator(task_frame, task, initial_state="pending")
                    task_indicator.pack(fill=tk.X, pady=3)
                    self.task_indicators[station_num][task] = icon_label
        
        # Configure grid weights
        for i in range(2):
            content_frame.rowconfigure(i, weight=1)
        for j in range(3):
            content_frame.columnconfigure(j, weight=1)
    
    def create_status_bar(self):
        # Create a status bar at the bottom
        status_frame = tk.Frame(self.root, bg="#E0E0E0", height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        
        # Status message
        status_label = tk.Label(status_frame, text="Ready", anchor="w", bg="#E0E0E0", font=("Arial", 10))
        status_label.pack(side=tk.LEFT, padx=15, pady=5)
        self.status_label = status_label
        
        # Batch time
        batch_time = tk.Label(status_frame, text="BATCH TIME: 00:00", bg="#4CAF50", fg="white", 
                             padx=10, pady=5, font=("Arial", 10, "bold"))
        batch_time.pack(side=tk.RIGHT, padx=15)
        self.batch_time = batch_time
        
        # System status
        status = tk.Label(status_frame, text="OFFLINE", bg="#FFD700", fg="#333", 
                         padx=10, pady=5, font=("Arial", 10, "bold"))
        status.pack(side=tk.RIGHT, padx=15)
        self.status = status

    def gather_com_ports_serial_numbers(self):
        """Accurately map serial numbers to COM ports based on station-channel alignment, skipping failed stations."""
        self.status_label.config(text="Mapping COM ports to serial numbers...")
        self.root.update()

        self.active_ports_serials = []

        # Station index to channel mapping (Station_1 = index 1 maps to Channel A, etc.)
        station_to_channel = {
            1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F"
            # Extend if more stations exist
        }

        valid_channels = ["A", "B", "C", "D", "E", "F", "G", "H"]

        # Step 1: Get all COM ports with their channels
        ports = serial.tools.list_ports.comports()
        channel_to_port = {}
        for port in ports:
            description_parts = port.description.split(' ')
            if len(description_parts) > 1:
                channel_name = description_parts[-2][-1]
                if channel_name in valid_channels:
                    channel_to_port[channel_name] = port.device

        # Step 2: Loop through station numbers and assign port if serial is present and station not failed
        for station_num, serial_var in self.station_numbers.items():
            if station_num in self.failed_stations:
                continue

            serial_number = serial_var.get().strip()
            if not serial_number:
                self.task_indicators[station_num]["Mapped with COM Port"].config(
                    text="⛔ Skipped", fg="#9E9E9E")
                continue

            channel = station_to_channel.get(station_num)
            com_port = channel_to_port.get(channel)

            if com_port:
                self.active_ports_serials.append((com_port, serial_number))

                if self.eng_var.get():
                    self.station_numbers[station_num].set(f"{com_port} - {serial_number}")
                    self.entries[station_num].config(
                        state="disabled", disabledbackground="#F5F5F5", disabledforeground="#333")

                self.task_indicators[station_num]["Mapped with COM Port"].config(
                    text="✓", fg="#4CAF50")
            else:
                self.task_indicators[station_num]["Mapped with COM Port"].config(
                    text="✕", fg="#F44336")

        logging.warning(f"Failed stations: {list(self.failed_stations)}")
        logging.info(f"Mapped COM Ports and Serial Numbers: {self.active_ports_serials}")
        logging.info(f"Skipped failed stations: {list(self.failed_stations)}")
        self.status_label.config(text="COM port mapping complete")

    @staticmethod
    def RF_test_perform(serial_port, test_router_address, network_address, network_channel, encryption_key, authentication_key, serial_number):
        try:
            logging.info(f"Starting RF test for Serial Number: {serial_number} on Port: {serial_port}")

            if getattr(sys, 'frozen', False):
                python_exe = os.path.join(os.path.dirname(sys.executable), "python.exe")
            else:
                python_exe = sys.executable

            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test.py")
            cmd = [
                python_exe,
                script_path,
                "--serial_port", serial_port,
                "--test_router_address", str(test_router_address),
                "--network_address", network_address,
                "--network_channel", str(network_channel),
                "--encryption_key", encryption_key,
                "--authentication_key", authentication_key,
                "--serial_number", serial_number
            ]
            logging.debug(f"RF Test command: {cmd}")

            try:
                result = subprocess.run(
                    cmd, capture_output=True, text=True, check=True,
                    cwd=os.path.dirname(script_path),
                    creationflags=subprocess.CREATE_NO_WINDOW
                )
                output = result.stderr
            except subprocess.CalledProcessError as e:
                output = e.stderr

            rx_rssi = re.search(r"RX RSSI: ([-\d]+ dBm)", output)
            tx_rssi = re.search(r"TX RSSI: ([-\d]+ dBm)", output)

            rx_rssi = rx_rssi.group(1) if rx_rssi else "N/A"
            tx_rssi = tx_rssi.group(1) if tx_rssi else "N/A"
            rf_ok = "RF is OK" in output
            flash_ok = "External Flash is OK" in output

            rf_status = "Pass" if rf_ok else "Fail"
            external_flash = "Pass" if flash_ok else "Fail"

            Action = f"Result - RX RSSI: {rx_rssi}, TX RSSI: {tx_rssi}, RF Status: {rf_status}, External Flash: {external_flash}"
            logging.info(f"Results - Serial Number: {serial_number}, RX RSSI: {rx_rssi}, TX RSSI: {tx_rssi}, RF Status: {rf_status}, External Flash: {external_flash}")

            # Save result to file
            result_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Detailed Test Result")
            os.makedirs(result_folder, exist_ok=True)
            result_file_name = f"{serial_number}_RF_Test_result_{rf_status.upper()}.txt"
            result_file_path = os.path.join(result_folder, result_file_name)

            mode = 'a' if os.path.exists(result_file_path) else 'w'
            with open(result_file_path, mode) as result_file:
                if mode == 'a':
                    result_file.write("\n--- Updated at {} ---\n".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                result_file.write(output)

            return serial_number, rf_status, external_flash, rx_rssi, tx_rssi, Action

        except Exception as e:
            logging.error(f"Unexpected error during RF Test for Serial Number {serial_number}: {e}")
            return serial_number, "Fail", "Fail", "N/A", "N/A", f"Error - {e}"

    def run_rf_test(self):
        """FIXED: Runs the RF test for all active stations with proper synchronization"""
        if not self.active_ports_serials:
            messagebox.showwarning("No Data", "Please enter serial numbers and ensure COM ports are available.")
            self.status_label.config(text="Please Re Run the test...")
            return

        self.status_label.config(text="Running RF tests...")
        
        # Initialize test completion tracking
        with self.test_completion_lock:
            self.completed_tests["rf_test"] = False
            self.station_test_results = {}
        
        # Update task indicators to show RF test is in progress - SKIP FAILED STATIONS
        for station_num, serial_var in self.station_numbers.items():
            if station_num in self.failed_stations:
                continue
                
            serial_number = serial_var.get().strip()
            if serial_number:
                self.task_indicators[station_num]["External Flash Test"].config(text="⏳", fg="#2196F3")
                self.task_indicators[station_num]["RSSI Test"].config(text="⏳", fg="#2196F3")
                ModernUI.update_status_indicator(self.status_indicators[station_num], "running")
        
        self.root.update()

        def perform_rf_test(com_port, serial_number):
            try:
                result = ManufacturingSuite.RF_test_perform(
                    serial_port=com_port,
                    test_router_address=self.TEST_ROUTER_ADDRESS,
                    network_address=self.NETWORK_ADDRESS,
                    network_channel=self.NETWORK_CHANNEL,
                    encryption_key=self.ENCRYPTION_KEY,
                    authentication_key=self.AUTHENTICATION_KEY,
                    serial_number=serial_number
                )

                serial, rf_status, flash_status, rx_rssi, tx_rssi, action = result

                # Save to Excel (thread-safe because each thread writes one row)
                sheet.append([
                    serial, rf_status, rx_rssi, tx_rssi, action,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
                workbook.save(file_name)

                return serial, rf_status, flash_status

            except Exception as e:
                logging.error(f"Error running RF test on {com_port} with serial {serial_number}: {e}")
                return serial_number, "Fail", "Fail"

        def execute_rf_tests():
            """FIXED: Runs all RF tests in parallel with proper completion tracking."""
            results = {}
            external_flash_results = {}
            
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future_to_test = {
                    executor.submit(perform_rf_test, com_port, serial_number): serial_number
                    for com_port, serial_number in self.active_ports_serials
                }
            
                # Wait for ALL futures to complete before updating UI
                for future in concurrent.futures.as_completed(future_to_test):
                    try:
                        serial_number, rf_status, external_flash = future.result()
                        results[serial_number] = rf_status
                        external_flash_results[serial_number] = external_flash
                        
                        # Store results for later status evaluation
                        with self.test_completion_lock:
                            self.station_test_results[serial_number] = {
                                'rf_status': rf_status,
                                'external_flash': external_flash
                            }
                            
                    except Exception as e:
                        logging.error(f"Exception in RF test execution: {e}")

            # CRITICAL FIX: Update UI only after ALL tests complete
            def update_ui_after_completion():
                # Update UI with RF test results - SKIP FAILED STATIONS
                for station_num, serial_var in self.station_numbers.items():
                    if station_num in self.failed_stations:
                        continue
                        
                    serial_number = serial_var.get().strip()
                    if self.eng_var.get() and " - " in serial_number:
                        serial_number = serial_number.split(" - ")[-1]  # Extract serial number if engineering mode is enabled

                    if serial_number in results:
                        rf_status = results[serial_number]
                        external_flash = external_flash_results[serial_number]
                        print(f"Station {station_num}: Serial {serial_number}, RF Status: {rf_status}, External Flash: {external_flash}")
                        logging.info(f"Station {station_num}: Serial {serial_number}, RF Status: {rf_status}, External Flash: {external_flash}")

                        self.task_indicators[station_num]["RSSI Test"].config(
                            text="✓" if rf_status == "Pass" else "✕",
                            fg="#4CAF50" if rf_status == "Pass" else "#F44336"
                        )

                        self.task_indicators[station_num]["External Flash Test"].config(
                            text="✓" if external_flash == "Pass" else "✕",
                            fg="#4CAF50" if external_flash == "Pass" else "#F44336"
                        )

                # Mark RF test as completed
                with self.test_completion_lock:
                    self.completed_tests["rf_test"] = True
                
                self.status_label.config(text="RF tests completed")
                
                # CRITICAL FIX: Update overall status only after UI updates are complete
                self.root.after(500, self.update_status_after_rf_tests)

            # Schedule UI update on main thread
            self.root.after(0, update_ui_after_completion)

        # Run RF tests in a separate thread
        threading.Thread(target=execute_rf_tests, daemon=True).start()

    def update_status_after_rf_tests(self):
        """NEW: Update overall station status after RF tests complete with proper synchronization"""
        with self.test_completion_lock:
            if not self.completed_tests.get("rf_test", False):
                # RF tests not yet completed, schedule another check
                self.root.after(100, self.update_status_after_rf_tests)
                return
        
        # Update station status indicators based on current test results
        for station_num, serial_var in self.station_numbers.items():
            if station_num in self.failed_stations:
                continue
                
            serial_number = serial_var.get().strip()
            if self.eng_var.get() and " - " in serial_number:
                serial_number = serial_number.split(" - ")[-1]
            
            if serial_number:
                # Check current status of all completed tests
                all_tests_passed = self.check_all_tests_status(station_num)
                
                # Only update to success/failure if we have definitive results
                if all_tests_passed is not None:
                    if all_tests_passed:
                        ModernUI.update_status_indicator(self.status_indicators[station_num], "success")
                    else:
                        ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                # If all_tests_passed is None, keep current "running" status

    def check_all_tests_status(self, station_num):
        """NEW: Check if all tests for a station have passed, failed, or are still pending"""
        test_tasks = [
            "Mapped with COM Port",
            "Test Firmware Download", 
            "External Flash Test",
            "RSSI Test",
            "Main Firmware Download",
            "Main Firmware Verification"
        ]
        
        passed_count = 0
        failed_count = 0
        pending_count = 0
        
        for task in test_tasks:
            if task in self.task_indicators[station_num]:
                status_text = self.task_indicators[station_num][task].cget("text")
                if status_text == "✓":
                    passed_count += 1
                elif status_text == "✕":
                    failed_count += 1
                else:
                    pending_count += 1
        
        # Return None if tests are still pending (don't update overall status yet)
        if pending_count > 0:
            return None
        
        # Return True only if ALL tests passed, False if any failed
        return failed_count == 0 and passed_count == len(test_tasks)
   
    def show_start_batch_dialog(self):
        # Dialog to confirm starting a new batch
        dialog = tk.Toplevel(self.root)
        dialog.title("Start New Batch")
        dialog.geometry("400x220")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="white")
        
        # Center the dialog
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (self.root.winfo_width() // 2) - (width // 2) + self.root.winfo_x()
        y = (self.root.winfo_height() // 2) - (height // 2) + self.root.winfo_y()
        dialog.geometry(f"+{x}+{y}")
        
        # Add a header
        header_frame = tk.Frame(dialog, bg="#2196F3", height=50)
        header_frame.pack(fill=tk.X)
        
        header_label = tk.Label(header_frame, text="Start New Batch", 
                               font=("Arial", 14, "bold"), bg="#2196F3", fg="white")
        header_label.pack(pady=10)
        
        # Dialog content
        content_frame = tk.Frame(dialog, bg="white")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Create info icon
        icon_canvas = tk.Canvas(content_frame, width=40, height=40, bg="white", highlightthickness=0)
        icon_canvas.pack(pady=5)
        
        # Draw info icon
        icon_canvas.create_oval(5, 5, 35, 35, fill="#E3F2FD", outline="#2196F3", width=2)
        icon_canvas.create_text(20, 20, text="i", font=("Arial", 18, "bold"), fill="#2196F3")
        
        message = tk.Label(content_frame, text="Are you sure you want to start a new batch?", 
                          bg="white", font=("Arial", 11))
        message.pack(pady=10)
        
        # Button frame
        button_frame = tk.Frame(content_frame, bg="white")
        button_frame.pack(pady=10)
        
        # Create buttons
        yes_btn_frame = ModernUI.create_modern_button(button_frame, "Yes", 
                                                    lambda: self.start_new_batch(dialog), 
                                                    width=100, height=30, bg="#4CAF50", hover_bg="#388E3C")
        yes_btn_frame.pack(side=tk.LEFT, padx=10)
        
        no_btn_frame = ModernUI.create_modern_button(button_frame, "No", 
                                                   dialog.destroy, 
                                                   width=100, height=30, bg="#F44336", hover_bg="#D32F2F")
        no_btn_frame.pack(side=tk.LEFT, padx=10)

    def start_new_batch(self, dialog):
        # Close dialog and start a new batch
        dialog.destroy()
        
        # Destroy the current root window
        self.root.destroy()
        
        # Create a new root window and initialize the GUI again
        new_root = tk.Tk()
        app = ManufacturingSuite(new_root)
        new_root.mainloop()
    
    def stop_test(self):
        """Stop all running tests and reset the status."""
        if self.timer_running:
            if self.timer_id:
                self.root.after_cancel(self.timer_id)
            self.timer_running = False
        
        self.status_label.config(text="Test stopped by user.")
        self.status.config(text="OFFLINE")
        self.batch_time.config(text="BATCH TIME: 00:00")

        for station_num, serial_var in self.station_numbers.items():
            self.station_numbers[station_num].set("")  # Clear the existing value
        
        # Reset status indicators
        for station_num in self.status_indicators:
            ModernUI.update_status_indicator(self.status_indicators[station_num], "pending")

        # Reset task indicators to default state
        for station_num, task_dict in self.task_indicators.items():
            for task, label in task_dict.items():
                label.config(text="⏱", fg="#757575")

        # Clear serial mappings and failed stations
        self.active_ports_serials.clear()
        self.failed_stations.clear()
        
        # Clear test completion tracking
        with self.test_completion_lock:
            self.completed_tests.clear()
            self.station_test_results.clear()

        # Unlock serial entries for re-entry
        for entry in self.entries.values():
            entry.config(state="normal")

    def FirmwareFlash(self, is_main_firmware=False):
        """Flash firmware to devices - MODIFIED to skip failed stations"""
        task_name = "Main Firmware Download" if is_main_firmware else "Test Firmware Download"
        Flashfile = "MainFirmwareFlash.py" if is_main_firmware else "FirmwareFlash.py"
        self.status_label.config(text=f"Running {task_name.lower()}...")
        
        count = 0
        for station_num, serial_var in self.station_numbers.items():
            # Skip failed stations
            if station_num in self.failed_stations:
                continue
                
            serial_number = serial_var.get().strip()  
            if serial_number:
                count += 1
                # Update task indicator to show firmware flash is in progress
                self.task_indicators[station_num][task_name].config(text="⏳", fg="#2196F3")
        
        self.root.update()

        folder_path = r'ConfigFiles'
        Config_file = 'GangConfig.cfg'

        try:
            process = subprocess.Popen(
                ["python", Flashfile, os.path.join(folder_path, Config_file)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )

            output, error = process.communicate()

            if error:
                logging.error(f"Firmware flash error: {error}")
                messagebox.showerror("Error", f"Failed to execute firmware flash: {error}")
                for station_num in range(1, 7):
                    if (station_num in self.task_indicators and 
                        station_num not in self.failed_stations and 
                        self.station_numbers[station_num].get().strip()):
                        self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
            else:
                verification_results = self.extract_verification_results(output)
                logging.info(f"Firmware flash results: {verification_results}")

                # Update task status based on flash result - SKIP FAILED STATIONS
                for station, result in verification_results.items():
                    try:
                        station_num = int(station.split("_")[1])
                        if station_num not in self.failed_stations:
                            if result == "OK":
                                self.task_indicators[station_num][task_name].config(text="✓", fg="#4CAF50")
                            else:
                                self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
                    except (ValueError, KeyError) as e:
                        logging.error(f"Error updating station {station} status: {e}")

        except Exception as e:
            logging.error(f"Failed to execute firmware flash: {e}")
            messagebox.showerror("Error", f"Failed to execute firmware flash: {e}")
            for station_num in range(1, 7):
                if (station_num in self.task_indicators and 
                    station_num not in self.failed_stations and 
                    self.station_numbers[station_num].get().strip()):
                    self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
    
    def extract_verification_results(self, output):
        """Extract verification results from firmware flash output"""
        pattern = r"#(\d+):\s+\.\.\s+(OK|failed)"
        matches = re.findall(pattern, output)
        return {f"Station_{num}": status for num, status in matches}
    
    def update_timer(self):
        """Update the batch time display"""
        if self.start_time and self.timer_running:
            elapsed_time = time.time() - self.start_time
            minutes = int(elapsed_time // 60)
            seconds = int(elapsed_time % 60)
            self.batch_time.config(text=f"BATCH TIME: {minutes:02d}:{seconds:02d}")
            self.timer_id = self.root.after(1000, self.update_timer)

    def check_main_firmware_version(self):
        """FIXED: Check firmware version with proper completion tracking"""
        task_name = "Main Firmware Verification"
        self.status_label.config(text=f"Verifying main firmware version...")
        
        # Initialize firmware test completion tracking
        with self.test_completion_lock:
            self.completed_tests["firmware_verification"] = False
        
        # Create the task indicator if it doesn't exist
        if task_name not in self.task_indicators[1]:
            for station_num in self.task_indicators:
                task_indicator, icon_label = ModernUI.create_task_indicator(
                    self.task_indicators[station_num]["Main Firmware Download"].master, 
                    task_name, 
                    initial_state="pending"
                )
                task_indicator.pack(fill=tk.X, pady=3)
                self.task_indicators[station_num][task_name] = icon_label
        
        # Update task indicators to show verification is in progress - SKIP FAILED STATIONS
        for station_num, serial_var in self.station_numbers.items():
            if station_num in self.failed_stations:
                continue
                
            serial_number = serial_var.get().strip()
            if serial_number:
                self.task_indicators[station_num][task_name].config(text="⏳", fg="#2196F3")
        
        self.root.update()
        
        # Path to firmware version config file
        firmware_config = os.path.join("ConfigFiles", "firmware_config.txt")
        
        # Check if config file exists, create it if it doesn't
        if not os.path.exists(firmware_config):
            os.makedirs(os.path.dirname(firmware_config), exist_ok=True)
            with open(firmware_config, 'w') as f:
                f.write("6.0.0.0")  # Default expected version
        
        def perform_firmware_check(com_port, serial_number):
            """Perform firmware version check for a single device."""
            try:
                # Using the correct function with the correct number of arguments
                serial_number, result, version, expected = check_firmware_version(
                    com_port, serial_number, firmware_config
                )
                logging.info(f"Firmware check completed for COM Port: {com_port}, Serial Number: {serial_number}")
                return serial_number, result, version, expected
            except Exception as e:
                logging.error(f"Error checking firmware on {com_port} with serial {serial_number}: {e}")
                return serial_number, False, None, None
        
        def execute_firmware_checks():
            """FIXED: Runs all firmware checks with proper completion tracking."""
            results = {}
            with concurrent.futures.ThreadPoolExecutor() as executor:
                future_to_check = {
                    executor.submit(perform_firmware_check, com_port, serial_number): serial_number
                    for com_port, serial_number in self.active_ports_serials
                }

                # Wait for ALL futures to complete
                for future in concurrent.futures.as_completed(future_to_check):
                    try:
                        serial_number, result, version, expected = future.result()
                        results[serial_number] = {
                            "result": result,
                            "version": version,
                            "expected": expected
                        }
                    except Exception as e:
                        logging.error(f"Exception in firmware check execution: {e}")

            # CRITICAL FIX: Update UI only after ALL checks complete
            def update_ui_after_firmware_completion():
                # Update UI with firmware check results - SKIP FAILED STATIONS
                for station_num, serial_var in self.station_numbers.items():
                    if station_num in self.failed_stations:
                        continue
                        
                    serial_number = serial_var.get().strip()
                    if self.eng_var.get() and " - " in serial_number:
                        serial_number = serial_number.split(" - ")[-1]  # Extract serial number if engineering mode is enabled

                    if serial_number in results:
                        result = results[serial_number]["result"]
                        version = results[serial_number]["version"]
                        expected = results[serial_number]["expected"]
                        
                        # Update task indicator
                        self.task_indicators[station_num][task_name].config(
                            text="✓" if result else "✕",
                            fg="#4CAF50" if result else "#F44336"
                        )
                        
                        # Add to Excel sheet
                        action = f"Firmware version: {version}, Expected: {expected}"
                        sheet.append([
                            serial_number, 
                            "Pass" if result else "Fail", 
                            "", 
                            "", 
                            action, 
                            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        ])
                        
                        # Save the Excel workbook
                        workbook.save(file_name)
                        logging.info(f"Firmware verification results for Serial Number: {serial_number}: {action}.")

                # Mark firmware verification as completed
                with self.test_completion_lock:
                    self.completed_tests["firmware_verification"] = True
                
                self.status_label.config(text="Firmware verification completed")
                
                # CRITICAL FIX: Update overall status after all tests complete
                self.root.after(500, self.update_status_after_tests)

            # Schedule UI update on main thread
            self.root.after(0, update_ui_after_firmware_completion)

        # Run firmware checks in a separate thread
        threading.Thread(target=execute_firmware_checks, daemon=True).start()

    def run_test(self):
        """Enhanced run test with auto-scanning integration - MODIFIED to handle failed stations"""

        for station_num, serial_var in self.station_numbers.items():
            self.station_numbers[station_num].set("")  # Clear the existing value
        
        # Reset status indicators
        for station_num in self.status_indicators:
            ModernUI.update_status_indicator(self.status_indicators[station_num], "pending")

        # Reset task indicators to default state
        for station_num, task_dict in self.task_indicators.items():
            for task, label in task_dict.items():
                label.config(text="⏱", fg="#757575")

        # Clear serial mappings and failed stations
        self.active_ports_serials.clear()
        self.failed_stations.clear()
        
        # Clear test completion tracking
        with self.test_completion_lock:
            self.completed_tests.clear()
            self.station_test_results.clear()

        # Unlock serial entries for re-entry
        for entry in self.entries.values():
            entry.config(state="normal")

        self.status.config(text="ONLINE")

        # Start timer
        self.start_time = time.time()
        self.timer_running = True
        self.update_timer()
        
        self.status_label.config(text="Test is running...")

        # Step 0: Auto-scan serial numbers if any are missing
        missing_serials = [station_num for station_num, serial_var in self.station_numbers.items() if not serial_var.get().strip()]

        if missing_serials:
            self.status_label.config(text="Missing serial numbers, starting auto-scan...")
            self.auto_scan_all_serials()
            time.sleep(2)  # Wait for scan to complete

        # Count active stations (excluding failed ones) and update status indicators
        count = 0
        for station_num, serial_var in self.station_numbers.items():
            if station_num in self.failed_stations:
                continue
                
            serial_number = serial_var.get().strip()  
            if serial_number:  
                count += 1
                ModernUI.update_status_indicator(self.status_indicators[station_num], "running")
        
        # Check if we have any valid stations to test
        if count == 0:
            self.status_label.config(text="No valid stations to test - all stations failed serial scan")
            logging.warning("No valid stations to test - all stations failed serial scan")
            return
        
        self.root.update()

        # Continue with existing test sequence for valid stations only...
        # Step 1: Map Serial Number with COM Ports
        self.gather_com_ports_serial_numbers()
        self.root.update()
        time.sleep(0.5)

        # Step 2: Test Firmware Download
        self.FirmwareFlash(is_main_firmware=False)
        self.root.update()  # Force UI update
        time.sleep(0.5)  # Allow user to see changes

        # Step 3: Run RF Tests
        self.run_rf_test()
        self.root.update()  # Force UI update
        time.sleep(15)  # Allow user to see changes

        # Step 4: Main Firmware Download
        self.FirmwareFlash(is_main_firmware=True)
        self.root.update()  # Force UI update
        time.sleep(0.5)  # Allow user to see changes
        
        # Step 5: Verify Main Firmware Version
        self.check_main_firmware_version()
        self.root.update()
        
        logging.info(f"Test completed in {time.time() - self.start_time:.2f} seconds")
        logging.info(f"Valid stations tested: {count}, Failed stations skipped: {len(self.failed_stations)}")

    def update_status_after_tests(self):
        """FIXED: Update station status after all tests complete with proper synchronization"""
        # Check if all tests are completed
        with self.test_completion_lock:
            rf_completed = self.completed_tests.get("rf_test", False)
            firmware_completed = self.completed_tests.get("firmware_verification", False)
            
            if not (rf_completed and firmware_completed):
                # Not all tests completed yet, schedule another check
                self.root.after(200, self.update_status_after_tests)
                return
        
        # All tests completed, now update final status
        for station_num, serial_var in self.station_numbers.items():
            # Skip already failed stations
            if station_num in self.failed_stations:
                continue
                
            serial_number = serial_var.get().strip()
            if serial_number:
                # Check if all tests passed for this station
                all_tests_passed = self.check_all_tests_status(station_num)
                
                # Update status indicator based on test results
                if all_tests_passed is True:
                    ModernUI.update_status_indicator(self.status_indicators[station_num], "success")
                elif all_tests_passed is False:
                    ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                # If all_tests_passed is None, keep current status (shouldn't happen at this point)
        
        # Stop timer
        self.timer_running = False
        self.status.config(text="OFFLINE")
        
        # Update status message with summary
        valid_stations = len([s for s in self.station_numbers.keys() if s not in self.failed_stations and self.station_numbers[s].get().strip()])
        failed_stations = len(self.failed_stations)
        
        if failed_stations > 0:
            self.status_label.config(text=f"Test completed - {valid_stations} stations tested, {failed_stations} stations failed/skipped")
        else:
            self.status_label.config(text="Test completed successfully.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ManufacturingSuite(root)
    root.mainloop()
