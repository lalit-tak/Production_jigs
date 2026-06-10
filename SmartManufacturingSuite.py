import tkinter as tk
from tkinter import ttk, messagebox
import serial
from pprint import pprint
import serial.tools.list_ports
import re
import subprocess
import time
import concurrent.futures
import threading
import pythoncom
import json
import logging
import sys
import os
import random
import psutil
import signal
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from PIL import Image, ImageTk, ImageDraw, ImageFont
import tkinter.simpledialog as simpledialog
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
import win32gui
import win32con
import win32com.client
import ast
from functools import partial

# Get current date for folder naming and filename
current_date = datetime.now().strftime("%Y-%m-%d")

# Setup directories

Test_Result = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Test Result Excel File")
Log_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Log files")
Ems_Result = os.path.join(os.path.dirname(os.path.abspath(__file__)), "EMS Test Result")

os.makedirs(Test_Result, exist_ok=True)
os.makedirs(Log_folder, exist_ok=True)
os.makedirs(Ems_Result, exist_ok=True)

log_folder = os.path.join(Log_folder, f"Log_file_{current_date}")
excel_folder = os.path.join(Test_Result, f"Test_Result_{current_date}")
ems_folder = os.path.join(Ems_Result, f"{current_date}")

os.makedirs(log_folder, exist_ok=True)
os.makedirs(excel_folder, exist_ok=True)
os.makedirs(ems_folder, exist_ok=True)

# Enhanced file naming with timestamp
timestamp_suffix = datetime.now().strftime("%H%M%S")
file_name = os.path.join(excel_folder, f"{current_date}_IMG_TEST_RESULT.xlsx")
detailed_log_file = os.path.join(excel_folder, f"{current_date}_DETAILED_TEST_LOG.xlsx")

def log_action(action, timestamp, serial_number=None):
    log_file_path = os.path.join(log_folder, "action_log.txt")
    with open(log_file_path, "a") as log_file:
        if serial_number:
            log_file.write(f"{timestamp} - Serial Number: {serial_number} - {action}\n")
        else:
            log_file.write(f"{timestamp} - {action}\n")

logging.basicConfig(
    filename=os.path.join(log_folder, f"{current_date}_program_log.txt"),
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

def insert_into_ems_text(data_row):
    if data_row[2] == "":
        return False
    else:
        result_file_name = f"{data_row[2]}_{data_row[-3].upper()}.txt"
        result_file_path = os.path.join(ems_folder, result_file_name)

        mode = 'a' if os.path.exists(result_file_path) else 'w'
        with open(result_file_path, mode) as f:
            if mode == 'a':
                f.write("\n--- Updated at {} ---\n".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            f.write("=" * 50 + "\n")
            f.write(f"Station Number :- {data_row[1]}\n")
            f.write(f"Serial Number :- {data_row[2]}\n")
            f.write(f"Start Time :- {data_row[3]}\n")
            f.write(f"End Time :- {data_row[3]}\n")
            f.write(f"Overall Result :- {data_row[-3]}\n")
            f.write(f"Total Time :- {data_row[-2]}\n")
            f.write(f"BOOT Mode Initialization :- {data_row[6]}\n")
            f.write(f"BOOT mode on :- {data_row[7]}\n")
            f.write(f"Buck Output :- {data_row[8]}\n")
            f.write(f"Buck Output Status :- {data_row[9]}\n")
            f.write(f"LDO Output :- {data_row[10]}\n")
            f.write(f"LDO Output Status :- {data_row[11]}\n")
            f.write(f"1V8 :- {data_row[14]}\n")
            f.write(f"1V8 Status :- {data_row[15]}\n")
            f.write(f"IMG SINK Programming :- {data_row[16]}\n")
            f.write(f"N58 Programming Initialization :- {data_row[17]}\n")
            f.write(f"Super Cap Output :- {data_row[12]}\n")
            f.write(f"Super Cap Output Status :- {data_row[13]}\n")
            f.write(f"BOOT Mode Off :- {data_row[18]}\n")
            f.write(f"Module Supply On :- {data_row[19]}\n")
            f.write(f"GW Core Ver :- {data_row[20]}\n")
            f.write(f"GW Core Ver Status :- {data_row[21]}\n")
            f.write(f"GW Main Ver :- {data_row[22]}\n")
            f.write(f"GW Main Ver Status :- {data_row[23]}\n")
            f.write(f"GW Sink App Ver :- {data_row[24]}\n")
            f.write(f"GW Sink App Ver Status :- {data_row[25]}\n")
            f.write(f"GW Modem Firmware Ver :- {data_row[26]}\n")
            f.write(f"GW Modem Firmware Ver Status :- {data_row[27]}\n")
            f.write(f"GW Meter Lib Ver :- {data_row[28]}\n")
            f.write(f"GW Meter Lib Ver Status :- {data_row[29]}\n")
            f.write(f"Version Check :- {data_row[30]}\n")
            f.write(f"Module Supply Off :- {data_row[31]}\n")
            f.write(f"SUPERCAP DISCHARGE :- {data_row[32]}\n")
            f.write(f"SUPERCAP DISCHARGED VOLTAGE :- {data_row[33]}\n")
            f.write(f"SUPERCAP DISCHARGED VOLTAGE Status :- {data_row[34]}\n")


            f.write("=" * 50 + "\n\n")
        return True

class ProcessManager:
    """Enhanced process management with forced termination capabilities"""
    
    def __init__(self):
        self.active_processes = {}
        self.process_lock = threading.Lock()
    
    def run_process_with_timeout(self, cmd, timeout, station_num, test_name):
        """Run a process with robust timeout and forced termination"""
        process = None
        start_time = time.time()
        
        try:
            logging.info(f"Station {station_num}: Starting process for {test_name} with {timeout}s timeout")
            
            # Start the process
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding='utf-8',
                creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
            )
            
            # Register the process
            with self.process_lock:
                self.active_processes[station_num] = {
                    'process': process,
                    'start_time': start_time,
                    'test_name': test_name,
                    'timeout': timeout
                }
            
            # Wait for completion with timeout
            try:
                stdout, stderr = process.communicate(timeout=timeout)
                return_code = process.returncode
                
                # Unregister the process
                with self.process_lock:
                    if station_num in self.active_processes:
                        del self.active_processes[station_num]
                
                logging.info(f"Station {station_num}: Process completed normally in {time.time() - start_time:.2f}s")
                return stdout, stderr, return_code, False  # False = not timed out
                
            except subprocess.TimeoutExpired:
                logging.error(f"Station {station_num}: Process timed out after {timeout}s, attempting forced termination")
                
                # Force kill the process and all its children
                self._force_kill_process_tree(process, station_num, test_name)
                
                # Unregister the process
                with self.process_lock:
                    if station_num in self.active_processes:
                        del self.active_processes[station_num]
                
                return "", f"Process timed out after {timeout} seconds", -1, True  # True = timed out
                
        except Exception as e:
            logging.error(f"Station {station_num}: Exception starting process: {e}")
            
            if process:
                self._force_kill_process_tree(process, station_num, test_name)
                
            # Unregister the process
            with self.process_lock:
                if station_num in self.active_processes:
                    del self.active_processes[station_num]
            
            return "", f"Exception: {str(e)}", -1, False
    
    def _force_kill_process_tree(self, process, station_num, test_name):
        """Forcefully kill a process and all its children"""
        try:
            # Get the process PID
            pid = process.pid
            logging.warning(f"Station {station_num}: Force killing process tree for {test_name} (PID: {pid})")
            
            # Try to get the process and its children using psutil
            try:
                parent = psutil.Process(pid)
                children = parent.children(recursive=True)
                
                # Kill all children first
                for child in children:
                    try:
                        logging.warning(f"Station {station_num}: Killing child process PID: {child.pid}")
                        child.kill()
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        pass
                
                # Kill the parent process
                try:
                    parent.kill()
                    logging.warning(f"Station {station_num}: Killed parent process PID: {pid}")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
                
                # Wait for processes to die
                gone, alive = psutil.wait_procs(children + [parent], timeout=3)
                
                # Force kill any remaining processes
                for p in alive:
                    try:
                        p.kill()
                        logging.warning(f"Station {station_num}: Force killed remaining process PID: {p.pid}")
                    except (psutil.NoSuchProcess, psutil.AccessDenied):
                        pass
                        
            except psutil.NoSuchProcess:
                logging.info(f"Station {station_num}: Process {pid} already terminated")
            
            # Fallback: use subprocess methods
            try:
                process.kill()
                process.wait(timeout=2)
            except:
                pass
            
            # Platform-specific force kill as last resort
            if sys.platform == "win32":
                try:
                    subprocess.run(['taskkill', '/F', '/T', '/PID', str(pid)], 
                                 capture_output=True, timeout=5)
                    logging.warning(f"Station {station_num}: Used taskkill for PID: {pid}")
                except:
                    pass
            else:
                try:
                    os.killpg(os.getpgid(pid), signal.SIGKILL)
                    logging.warning(f"Station {station_num}: Used killpg for PID: {pid}")
                except:
                    pass
                    
        except Exception as e:
            logging.error(f"Station {station_num}: Error in force kill: {e}")
    
    def kill_station_processes(self, station_num):
        """Kill all processes for a specific station"""
        with self.process_lock:
            if station_num in self.active_processes:
                process_info = self.active_processes[station_num]
                process = process_info['process']
                test_name = process_info['test_name']
                
                logging.warning(f"Station {station_num}: Killing process for {test_name}")
                self._force_kill_process_tree(process, station_num, test_name)
                del self.active_processes[station_num]
    
    def get_active_processes(self):
        """Get information about active processes"""
        with self.process_lock:
            return dict(self.active_processes)
    
    def cleanup_all_processes(self):
        """Clean up all active processes"""
        with self.process_lock:
            for station_num, process_info in list(self.active_processes.items()):
                process = process_info['process']
                test_name = process_info['test_name']
                logging.warning(f"Station {station_num}: Cleaning up process for {test_name}")
                self._force_kill_process_tree(process, station_num, test_name)
            self.active_processes.clear()

class ExcelLogger:
    """Enhanced Excel logging with detailed test data and formatting"""
    
    def __init__(self, main_file, detailed_file):
        self.main_file = main_file
        self.detailed_file = detailed_file
        self.lock = threading.Lock()
        
        # Color schemes for Excel formatting
        self.colors = {
            'pass': PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
            'fail': PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid'),
            'header': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
            'warning': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        }
        
        self.fonts = {
            'header': Font(bold=True, size=11),
            'normal': Font(size=10),
            'bold': Font(bold=True, size=10)
        }
        
        self.initialize_excel_files()
    
    def initialize_excel_files(self):
        """Initialize both main and detailed Excel files with headers"""
        with self.lock:
            # Initialize main results file
            if not os.path.exists(self.main_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "Test Results Summary"
                
                headers = [
                    "Test Session", "Station", "PCBA Serial Number", "Test Start Time", "Test End Time",
                    "Mapped with COM Port", "BOOT Mode Initialization", "BOOT mode on", "Buck Output", "Buck Output Status", "LDO Output", "LDO Output Status", "1V8", "1V8 Status", "IMG SINK Programming", "N58 Programming Initialization", "Super Cap Voltage", "Super Cap Voltage Status", "BOOT Mode Off", "Module Supply On", "GW Core Ver", "GW Core Ver Status", "GW Main Ver", "GW Main Ver Status", "GW Sink App Ver", "GW Sink App Ver Status", "GW Modem Firmware Ver", "GW Modem Firmware Ver Status","GW Meter Lib Ver","GW Meter Lib Ver Status", "Version Check",   "Module Supply Off", "SUPERCAP DISCHARGE", "SUPERCAP DISCHARGED VOLTAGE", "SUPERCAP DISCHARGED VOLTAGE Status",
                    "Overall Status", "Total Test Time (s)", "Timestamp"
                ]
                
                ws.append(headers)
                self._format_header_row(ws, len(headers))
                wb.save(self.main_file)
            
            # Initialize detailed log file
            if not os.path.exists(self.detailed_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "IMG Detailed Test Logs"
                
                headers = [
                    "Test Session", "Station", "Serial Number", "Test Name", "Test Start Time",
                    "Test End Time", "Test Duration (s)", "Command Sent", "Raw Response", 
                     "Status", "Error Message", "Retry Count", "Timestamp"
                ]
                
                ws.append(headers)
                self._format_header_row(ws, len(headers))
                wb.save(self.detailed_file)
    
    def _format_header_row(self, worksheet, num_columns):
        """Format the header row with styling"""
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = self.colors['header']
            cell.font = self.fonts['header']
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def log_test_start(self, session_id, station, serial_number):
        """Log when a test session starts"""
        with self.lock:
            try:
                wb = load_workbook(self.detailed_file)
                ws = wb.active
                
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([
                    session_id, station, serial_number, "TEST_SESSION_START", timestamp,
                    "", "", "", "", "STARTED", "", "", timestamp
                ])
                
                wb.save(self.detailed_file)
                logging.info(f"Logged test start for Station {station}, Serial: {serial_number}")
            except Exception as e:
                logging.error(f"Error logging test start: {e}")
    
    def log_individual_test(self, session_id, station, serial_number, test_name, 
                          start_time, end_time, command, response, status, error_msg="", retry_count=0):
        """Log individual test details"""
        with self.lock:
            try:
                wb = load_workbook(self.detailed_file)
                ws = wb.active
                
                duration = (end_time - start_time).total_seconds() if end_time and start_time else 0
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                row_data = [
                    session_id, station, serial_number, test_name,
                    start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else "",
                    end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "",
                    f"{duration:.2f}", str(command)[:100], str(response)[:200], status, error_msg, retry_count, timestamp
                ]
                
                ws.append(row_data)
                
                # Apply color formatting based on status
                row_num = ws.max_row
                status_cell = ws.cell(row=row_num, column=11)  # Status column
                if status == "PASS":
                    status_cell.fill = self.colors['pass']
                elif status == "FAIL":
                    status_cell.fill = self.colors['fail']
                elif status == "ERROR":
                    status_cell.fill = self.colors['warning']
                
                wb.save(self.detailed_file)
                logging.debug(f"Logged individual test: {test_name} for {serial_number}")
            except Exception as e:
                logging.error(f"Error logging individual test: {e}")
    
    def log_final_results(self, session_id, station, serial_number, test_results, 
                         start_time, end_time):
        """Log final test results to main file with proper value extraction"""
        with self.lock:
            try:
                wb = load_workbook(self.main_file)
                ws = wb.active
                overall_status = test_results.get('overall_result').get('status', 'FAIL')
                total_time = (end_time - start_time).total_seconds() if end_time and start_time else 0
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Extract individual test results with proper value handling
                COM_PORT_MAPPING = test_results.get('COM_PORT_MAPPING', {})
                STATIC_DO0_ON = test_results.get('BOOT Mode Initialization', {})
                BOOT_Mode = test_results.get('BOOT MODE ON', {})
                Instant_AI_Buck_Output = test_results.get("Buck Output", {})
                Instant_AI_LDO_output = test_results.get("LDO output", {})
                Instant_AI_Super_Cap_Voltage = test_results.get("Super Cap Output", {})
                Instant_AI_1V8 = test_results.get('1V8', {})
                IMG_SINK_Programming = test_results.get('IMG SINK Programming', {})
                VERSION_CHECK = test_results.get('Firmware Version Check', {})
                GW_CORE_VER = test_results.get('GW_CORE_VER', {})
                GW_MAIN_VER = test_results.get('GW_MAIN_VER', {})
                GW_SINK_APP_VER = test_results.get('GW_SINK_APP_VER', {})
                GW_MODEM_FIRMWARE_VER = test_results.get('GW_MODEM_FIRMWARE_VER', {})
                GW_METER_LIB_VER = test_results.get('GW_METER_LIB_VER', {})
                N58_Programming_Initialization = test_results.get('N58 Programming Initialization', {})
                BOOT_MODE_OFF = test_results.get('BOOT MODE OFF', {})
                MODULE_SUPPLY_ON = test_results.get('MODULE SUPPLY ON', {})
                # sim1_cellular_test_status = test_results.get('sim1_cellular_test_status', {})
                # sim2_cellular_test_status = test_results.get('sim2_cellular_test_status', {})
                # gpio_test_status = test_results.get('gpio_test_status', {})
                # external_flash_test_status = test_results.get('external_flash_test_status', {})
                # rx_rssi_value = test_results.get('rx_rssi_value', {})
                # tx_rssi_value = test_results.get('tx_rssi_value', {})
                # rf_test_status = test_results.get('rf_test_status', {})
                MODULE_SUPPLY_OFF = test_results.get('MODULE SUPPLY OFF', {})
                SUPERCAP_DISCHARGE = test_results.get('SUPERCAP DISCHARGE', {})
                SUPERCAP_DISCHARGE_VOLTAGE = test_results.get('SUPERCAP DISCHARGED VOLTAGE', {})
                overall_station_status = test_results.get('overall_result', {})
                

                row_data = [
                    session_id, station, serial_number,
                    start_time.strftime("%Y-%m-%d %H:%M:%S") if start_time else "",
                    end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time else "",
                    COM_PORT_MAPPING.get('status', 'FAIL'),
                    STATIC_DO0_ON.get('status','FAIL'),
                    BOOT_Mode.get('status','FAIL'),
                    Instant_AI_Buck_Output.get("value", ''),
                    Instant_AI_Buck_Output.get("status", "FAIL"),
                    Instant_AI_LDO_output.get("value", ''),
                    Instant_AI_LDO_output.get("status", "FAIL"),
                    Instant_AI_1V8.get('value', ''),
                    Instant_AI_1V8.get("status", "FAIL"),
                    IMG_SINK_Programming.get('status','FAIL'),
                    N58_Programming_Initialization.get('status','FAIL'),
                    Instant_AI_Super_Cap_Voltage.get("value", ''),
                    Instant_AI_Super_Cap_Voltage.get("status", "FAIL"),
                    BOOT_MODE_OFF.get('status','FAIL'),
                    MODULE_SUPPLY_ON.get('status','FAIL'),
                    GW_CORE_VER.get('value', ""),
                    GW_CORE_VER.get('status', ""),
                    GW_MAIN_VER.get('value', ""),
                    GW_MAIN_VER.get('status', ""),
                    GW_SINK_APP_VER.get('value', ""),
                    GW_SINK_APP_VER.get('status', ""),
                    GW_MODEM_FIRMWARE_VER.get('value', ""),
                    GW_MODEM_FIRMWARE_VER.get('status', ""),
                    GW_METER_LIB_VER.get('value', ""),
                    GW_METER_LIB_VER.get('status', ""),
                    VERSION_CHECK.get('status', ""),
                    # sim1_cellular_test_status.get('status','FAIL'),
                    # sim2_cellular_test_status.get('status','FAIL'),
                    # gpio_test_status.get('status','FAIL'),
                    # external_flash_test_status.get('status','FAIL'),
                    # rx_rssi_value.get('value',' '),
                    # tx_rssi_value.get('value',' '),
                    # rf_test_status.get('status','FAIL'),
                    MODULE_SUPPLY_OFF.get('status','FAIL'),
                    SUPERCAP_DISCHARGE.get('status','FAIL'),
                    SUPERCAP_DISCHARGE_VOLTAGE.get('value', ''),
                    SUPERCAP_DISCHARGE_VOLTAGE.get('status', 'FAIL'),
                    overall_station_status.get('status', 'FAIL')
                ]
                # row_data.append(overall_status)
                row_data.append(f"{total_time:.2f}")
                row_data.append(timestamp)


                
                ws.append(row_data)
                insert_into_ems_text(row_data)
                # Apply formatting to the entire row based on overall status
                row_num = ws.max_row
                fill_color = self.colors['pass'] if overall_status == 'PASS' else self.colors['fail']
                
                for col in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    if col == len(row_data) - 2:  # Overall status column
                        cell.fill = fill_color
                        cell.font = self.fonts['bold']
                
                wb.save(self.main_file)
                logging.info(f"Logged final results for Station {station}, Serial: {serial_number}, Status: {overall_status}")
            except Exception as e:
                logging.error(f"Error logging final results: {e}")
    
    def _extract_memory_value(self, memory_result, mem_type):
        """Extract memory test values (EEPROM/FLASH)"""
        if not memory_result:
            return "N/A"
        
        value = memory_result.get('value', {})
        if isinstance(value, dict):
            return value.get(mem_type, 'N/A')
        return "N/A"
    
    def _extract_current_value(self, current_result, phase):
        """Extract current values for specific phase"""
        if not current_result:
            return "N/A"
        
        value = current_result.get('value', {})
        if isinstance(value, dict) and phase in value:
            curr_val = value[phase]
            if isinstance(curr_val, (int, float)):
                return f"{curr_val:.3f}"
        return "N/A"
    
    def _extract_voltage_value(self, voltage_result, phase):
        """Extract voltage values for specific phase"""
        if not voltage_result:
            return "N/A"
        
        value = voltage_result.get('value', {})
        if isinstance(value, dict) and phase in value:
            volt_val = value[phase]
            if isinstance(volt_val, (int, float)):
                return f"{volt_val:.3f}"
        return "N/A"

class SerialScanner:
    """Enhanced serial scanning functionality with retry mechanism and station isolation"""
    
    def __init__(self, port_config, baud_rate=9600, trigger_command=b'*T', timeout=3, max_retries=1):
        self.port_config = port_config
        self.baud_rate = baud_rate
        self.trigger_command = trigger_command
        self.timeout = timeout
        self.max_retries = max_retries
        self.scan_results = {}
        self.scan_lock = threading.Lock()
        self.retry_stations = set()
    
    def scan_single_port(self, station_num, port, is_retry=False):
        """Scan a single COM port for serial number with improved error handling"""
        retry_text = " (RETRY)" if is_retry else ""

        try:
            logging.info(f"[Station {station_num}] Scanning port {port}{retry_text}")
            with serial.Serial(port, self.baud_rate, timeout=self.timeout) as ser:
                ser.write(self.trigger_command)
                logging.debug(f"[Station {station_num}] Sent trigger command to {port}{retry_text}")
                
                time.sleep(1)
                scanned_data = ser.readline().decode('utf-8', errors='ignore').strip()
                # scanned_data = ser.read_all()

                # scanned_data = "ABC-123E2"

                with self.scan_lock:
                    if scanned_data:
                        self.scan_results[station_num] = {
                            'serial': scanned_data,
                            'port': port,
                            'status': 'success',
                            'retry_attempt': is_retry
                        }
                        self.retry_stations.discard(station_num)
                        logging.info(f"[Station {station_num}] Successfully scanned{retry_text}: {scanned_data}")
                    else:
                        self.scan_results[station_num] = {
                            'serial': None,
                            'port': port,
                            'status': 'no_data',
                            'retry_attempt': is_retry
                        }
                        if not is_retry:
                            self.retry_stations.add(station_num)
                        logging.warning(f"[Station {station_num}] No data received from {port}{retry_text}")
                        
        except serial.SerialException as e:
            with self.scan_lock:
                self.scan_results[station_num] = {
                    'serial': None,
                    'port': port,
                    'status': 'error',
                    'error': str(e),
                    'retry_attempt': is_retry
                }
                if not is_retry:
                    self.retry_stations.add(station_num)
            logging.error(f"[Station {station_num}] Serial error on {port}{retry_text}: {e}")
        except Exception as e:
            with self.scan_lock:
                self.scan_results[station_num] = {
                    'serial': None,
                    'port': port,
                    'status': 'error',
                    'error': str(e),
                    'retry_attempt': is_retry
                }
                if not is_retry:
                    self.retry_stations.add(station_num)
            logging.error(f"[Station {station_num}] Unexpected error on {port}{retry_text}: {e}")
    
    def scan_all_configured_ports(self):
        """Scan all configured COM ports in parallel with retry mechanism and station isolation"""
        self.scan_results.clear()
        self.retry_stations.clear()
        
        logging.info("Starting initial scan of all configured ports")
        self._perform_scan_batch(is_retry=False)
        if self.retry_stations and self.max_retries > 0:
            logging.info(f"Retrying failed stations: {list(self.retry_stations)}")
            time.sleep(2)
            self._perform_scan_batch(is_retry=True, stations_to_scan=self.retry_stations.copy())
        
        return self.scan_results.copy()
    
    def _perform_scan_batch(self, is_retry=False, stations_to_scan=None):
        """Perform a batch of scans (either initial or retry) with station isolation"""
        threads = []
        
        if stations_to_scan is None:
            stations_to_scan = self.port_config.keys()
        
        for station_num in stations_to_scan:
            port = self.port_config.get(station_num)
            if port:
                thread = threading.Thread(
                    target=self.scan_single_port, 
                    args=(station_num, port, is_retry),
                    daemon=True
                )
                threads.append(thread)
                thread.start()
        
        for thread in threads:
            thread.join()
        
        return self.scan_results.copy()

class ToggleSwitch(tk.Canvas):
    def __init__(self, parent, width=60, height=30, 
                 on_color="green", off_color="#B0BEC5", 
                 disabled_color="#CFD8DC", **kwargs):
        super().__init__(parent, width=width, height=height, 
                         bg=parent["bg"], highlightthickness=0, **kwargs)

        self.width = width
        self.height = height
        self.on_color = on_color
        self.off_color = off_color
        self.disabled_color = disabled_color
        self.is_on = False
        self.state = "normal"   # can be "normal" or "disabled"

        # Draw background
        self.bg_rect = self.create_oval(2, 2, height-2, height-2, fill=self.off_color, outline=self.off_color)
        self.bg_rect2 = self.create_oval(width-height, 2, width-2, height-2, fill=self.off_color, outline=self.off_color)
        self.bg_rect3 = self.create_rectangle(height//2, 2, width-height//2, height-2, fill=self.off_color, outline=self.off_color)

        # Knob
        self.knob = self.create_oval(2, 2, height-2, height-2, fill="white", outline="gray")

        # Bind click events to all parts
        for item in (self.bg_rect, self.bg_rect2, self.bg_rect3, self.knob):
            self.tag_bind(item, "<Button-1>", self.toggle)

    def toggle(self, event=None):
        if self.state == "disabled":
            return
        self.is_on = not self.is_on
        self.update_ui()

    def update_ui(self):
        # Update background color
        if self.state == "disabled":
            color = self.disabled_color
            knob_color = "#ECEFF1"
        else:
            color = self.on_color if self.is_on else self.off_color
            knob_color = "white"

        for item in (self.bg_rect, self.bg_rect2, self.bg_rect3):
            self.itemconfig(item, fill=color, outline=color)

        self.itemconfig(self.knob, fill=knob_color, outline="gray")

        # Move knob
        if self.is_on:
            self.coords(self.knob, self.width - self.height + 2, 2, self.width - 2, self.height - 2)
        else:
            self.coords(self.knob, 2, 2, self.height - 2, self.height - 2)

        self.update_idletasks()

    def turn_off(self):
        self.is_on = False
        self.update_ui()

    def enable(self):
        self.state = "normal"
        self.update_ui()

    def disable(self):
        self.state = "disabled"
        self.update_ui()

class ModernUI:
    """Helper class for modern UI elements"""
    
    @staticmethod
    def create_rounded_rectangle(canvas, x1, y1, x2, y2, radius=25, **kwargs):
        """Draw a rounded rectangle on a canvas"""
        points = [
            x1+radius, y1, x2-radius, y1, x2, y1, x2, y1+radius,
            x2, y2-radius, x2, y2, x2-radius, y2, x1+radius, y2,
            x1, y2, x1, y2-radius, x1, y1+radius, x1, y1
        ]
        return canvas.create_polygon(points, **kwargs, smooth=True)
    
    @staticmethod
    def create_status_indicator(parent, size=60, initial_state="pending"):
        """Create a modern status indicator"""
        canvas = tk.Canvas(parent, width=size, height=size, bg=parent["bg"], highlightthickness=0)
        
        states = {
            "pending": {"fill": "#E0E0E0", "outline": "#BDBDBD", "symbol": "⏱", "symbol_color": "#757575"},
            "scanning": {"fill": "#FFE082", "outline": "#FFC107", "symbol": "📡", "symbol_color": "#F57F17"},
            "retrying": {"fill": "#FFCC80", "outline": "#FF9800", "symbol": "🔄", "symbol_color": "#E65100"},
            "running": {"fill": "#90CAF9", "outline": "#2196F3", "symbol": "⏳", "symbol_color": "#0D47A1"},
            "success": {"fill": "#A5D6A7", "outline": "#4CAF50", "symbol": "✓", "symbol_color": "#1B5E20"},
            "failure": {"fill": "#EF9A9A", "outline": "#F44336", "symbol": "✕", "symbol_color": "#B71C1C"}
        }
        
        canvas.create_oval(3, 3, size-1, size-1, fill="#F5F5F5", outline="#E0E0E0", width=2)
        
        state_config = states.get(initial_state, states["pending"])
        inner_size = 4
        canvas.create_oval(inner_size, inner_size, size-inner_size, size-inner_size, 
                          fill=state_config["fill"], outline=state_config["outline"], width=2)
        
        canvas.create_text(size/2, size/2, text=state_config["symbol"], 
                          fill=state_config["symbol_color"], font=("Arial", int(size/2), "bold"))
        
        return canvas

    @staticmethod
    def update_status_indicator(canvas, state):
        """Update the status indicator to a new state"""
        size = canvas.winfo_width()
        if size < 10:
            size = 60
        
        canvas.delete("all")
        
        states = {
            "pending": {"fill": "#E0E0E0", "outline": "#BDBDBD", "symbol": "⏱", "symbol_color": "#757575"},
            "scanning": {"fill": "#FFE082", "outline": "#FFC107", "symbol": "📡", "symbol_color": "#F57F17"},
            "retrying": {"fill": "#FFCC80", "outline": "#FF9800", "symbol": "🔄", "symbol_color": "#E65100"},
            "running": {"fill": "#90CAF9", "outline": "#2196F3", "symbol": "⏳", "symbol_color": "#0D47A1"},
            "success": {"fill": "#4CAF50", "outline": "#2E7D32", "symbol": "✓", "symbol_color": "#FFFFFF"},
            "failure": {"fill": "#F44336", "outline": "#B71C1C", "symbol": "✕", "symbol_color": "#FFFFFF"}
        }
        
        state_config = states.get(state, states["pending"])
        
        gradient_colors = ["#D0D0D0", "#A0A0A0"]
        steps = 10
        step_size = (size - 4) / (2 * steps)
        
        for i in range(steps):
            color = gradient_colors[i % 2]
            canvas.create_oval(2 + i * step_size, 2 + i * step_size, size - 2 - i * step_size, size - 2 - i * step_size,
                              outline=color, width=1)
        
        gradient_offset = 4
        canvas.create_oval(gradient_offset, gradient_offset, size - gradient_offset, size - gradient_offset,
                          fill=state_config["fill"], outline=state_config["outline"], width=2)
        
        highlight_offset = 8
        canvas.create_arc(highlight_offset, highlight_offset, size - highlight_offset, size - highlight_offset,
                          start=30, extent=120, style="arc", outline="#FFFFFF", width=2)
        
        shadow_offset = 1
        canvas.create_oval(gradient_offset + shadow_offset, gradient_offset + shadow_offset,
                          size - gradient_offset - shadow_offset, size - gradient_offset - shadow_offset,
                          outline="#808080", width=1)
        
        canvas.create_text(size / 2, size / 2, text=state_config["symbol"],
                          fill=state_config["symbol_color"], font=("Arial", int(size / 2.5), "bold"))

    @staticmethod
    def create_modern_button(parent, text, command, width=120, height=40, bg="#4CAF50", hover_bg="#388E3C"):
        def on_enter(e):
            if button._enabled:
                button["bg"] = hover_bg
                button["relief"] = "raised"
        
        def on_leave(e):
            if button._enabled:
                button["bg"] = bg
                button["relief"] = "ridge"
        
        def on_press(e):
            if button._enabled:
                button["relief"] = "sunken"
        
        def on_release(e):
            if button._enabled:
                button["relief"] = "raised"
                command()   # only fire if enabled ✅

        button = tk.Label(
            parent,
            text=text,
            width=width, height=height,
            bg=bg, fg="white",
            font=("Arial", 12, "bold"),
            bd=3, relief="ridge",
            padx=5, pady=5,
            cursor="hand2"
        )

        # Enable/disable state flag
        button._enabled = True
        button._bg = bg
        button._hover_bg = hover_bg

        # Bind events
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)
        button.bind("<ButtonPress-1>", on_press)
        button.bind("<ButtonRelease-1>", on_release)

        # Custom methods
        def enable():
            button._enabled = True
            button.config(fg="white", bg=button._bg, cursor="hand2")

        def disable():
            button._enabled = False
            button.config(fg="gray", bg="#B0BEC5", cursor="arrow")

        button.enable = enable
        button.disable = disable

        button.pack_propagate(False)
        return button

    @staticmethod
    def create_task_indicator(parent, text, initial_state="pending"):
        """Create a task indicator with icon and text"""
        frame = tk.Frame(parent, bg=parent["bg"])
        
        states = {
            "pending": {"symbol": "⏱", "color": "#757575"},
            "running": {"symbol": "⏳", "color": "#2196F3"},
            "success": {"symbol": "✓", "color": "#4CAF50"},
            "failure": {"symbol": "✕", "color": "#F44336"}
        }
        
        state_config = states.get(initial_state, states["pending"])
        
        icon_label = tk.Label(frame, text=state_config["symbol"], fg=state_config["color"], 
                             bg=parent["bg"], font=("Arial", 12, "bold"))
        icon_label.pack(side=tk.LEFT, padx=(0, 5))
        
        text_label = tk.Label(frame, text=text, anchor="w", bg=parent["bg"], font=("Arial", 9))
        text_label.pack(side=tk.LEFT, fill=tk.X)
        
        return frame, icon_label

class ManufacturingSuite:
    def __init__(self, root):
        self.root = root
        self.root.title("Smart Manufacturing Suite : IMG- [SMS] [Version 1.0.0.0, 18-JULY-25]")
        self.root.geometry("1350x700")
        self.root.configure(bg="#F5F5F5")
        
        # Initialize process manager
        self.process_manager = ProcessManager()
        
        self.root.attributes('-fullscreen', True)
        self.root.bind("<Escape>", self.exit_fullscreen)
        
        # Initialize Excel logger
        self.excel_logger = ExcelLogger(file_name, detailed_log_file)
        
        # Generate unique session ID for this test run
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")

        self.baudrate = 19200
        
        # firmware version from text file
        self.firmware_version = None

        # Load configuration for COM ports mapping
        self.scanner_ports = {}
        self.optical_ports = {}
        self.dec_ports = {}
        self.load_com_port_config("config.txt")

        # Load Version Details
        self.core_ver = None
        self.main_ver = None
        self.sink_app_ver = None
        self.modem_firmware_ver = None
        self.meter_lib_ver = None
        self.load_version_config('version_config.json')

        # Min Max Voltage of Instant Ai
        self.min_buck_voltage = 3.65
        self.max_buck_voltage = 4.2
        self.min_ldo_output = 3.2
        self.max_ldo_output = 3.45
        self.min_super_cap_voltage = 0.15
        self.max_super_cap_voltage = 2.75
        self.min_v8_voltage = 1.7
        self.max_v8_voltage = 2.0
        self.min_super_cap_discharge_voltage = 0
        self.max_super_cap_discharge_voltage = 1.2
        
        # Initialize serial scanner with retry capability
        self.serial_scanner = SerialScanner(self.scanner_ports, max_retries=1)
        
        self.active_ports_serials = []
        self.task_labels = {}
        self.station_numbers = {}
        self.entries = {}
        self.scanned_serials = set()
        self.eng_var = tk.BooleanVar()
        self.status_indicators = {}
        self.task_indicators = {}
        
        # Track failed stations
        self.failed_stations = set()
        
        # Enhanced test tracking with timing - per station tracking
        self.test_completion_lock = threading.Lock()
        self.station_completed_tests = {}  # Track completion per station
        self.completed_tests = {
            "Mapped with COM Port": False,
            "BOOT Mode Initialization": False,
            "BOOT MODE ON": False,
            "Buck Output": False,
            "IMG SINK Programming": False,
            "N58 Programming Initialization": False,
            "Super Cap Output": False,
            "BOOT MODE OFF": False,
            "MODULE SUPPLY ON": False,
            "Firmware Version Check": False,
            "MODULE SUPPLY OFF": False,
            "SUPERCAP DISCHARGE": False,
            "SUPERCAP DISCHARGED VOLTAGE": False,
            "Module Supply Disconnected": False
        }
        self.station_test_results = {}
        self.station_test_times = {}  # Track start/end times per station
        
        # Create the UI components
        self.create_header()
        self.create_main_content()
        self.create_status_bar()
        
        # Timer variables
        self.start_time = None
        self.timer_running = False
        self.timer_id = None

        # Cycle and pass counters
        self.total_cycles = 0
        self.total_passed_stations = 0
        self.is_break = False
        
        # Station-specific test threads and control flags
        self.station_threads = {}
        self.station_stop_flags = {}
        
        # Flag to track if any test has failed in the programming phase
        self.programming_phase_failed = False
        
        # Start process monitoring thread
        self.start_process_monitor()
        
    def load_version_config(self, file_name):
        data_dict = {}
        with open(file_name, "r") as file:
            data_dict = json.load(file)
        self.core_ver = data_dict["core_ver"]
        self.main_ver = data_dict["main_ver"]
        self.sink_app_ver = data_dict["sink_app_ver"]
        self.modem_firmware_ver = data_dict["modem_firmware_ver"]
        self.meter_lib_ver = data_dict["meter_lib_ver"]

    def start_process_monitor(self):
        """Start a background thread to monitor processes"""
        def monitor_processes():
            while True:
                try:
                    active_processes = self.process_manager.get_active_processes()
                    current_time = time.time()
                    
                    for station_num, process_info in active_processes.items():
                        start_time = process_info['start_time']
                        timeout = process_info['timeout']
                        test_name = process_info['test_name']
                        
                        # Check if process has been running too long (timeout + 10 seconds grace period)
                        if current_time - start_time > timeout + 10:
                            logging.warning(f"Station {station_num}: Process {test_name} exceeded timeout by 10s, force killing")
                            self.process_manager.kill_station_processes(station_num)
                            
                            # Mark station as failed
                            self.failed_stations.add(station_num)
                            self.station_stop_flags[station_num] = True
                    
                    time.sleep(5)  # Check every 5 seconds
                except Exception as e:
                    logging.error(f"Error in process monitor: {e}")
                    time.sleep(5)
        
        monitor_thread = threading.Thread(target=monitor_processes, daemon=True)
        monitor_thread.start()

    def load_com_port_config(self, path):
        """Load COM port configuration from config.txt"""
        if os.path.exists(path):
            try:
                with open(path, "r") as f:
                    for line in f:
                        line = line.strip()
                        if "=" in line and not line.startswith("#"):
                            key, value = line.split("=", 1)
                            key = key.strip()
                            value = value.strip()
                            if key == "SCANNER_COM1":
                                self.scanner_ports[1] = value
                            elif key == "SCANNER_COM2":
                                self.scanner_ports[2] = value
                            elif key == "SCANNER_COM3":
                                self.scanner_ports[3] = value
                            elif key == "SCANNER_COM4":
                                self.scanner_ports[4] = value
                            elif key == "OPTICAL_COM1":
                                self.optical_ports[1] = value
                            elif key == "OPTICAL_COM2":
                                self.optical_ports[2] = value
                            elif key == "OPTICAL_COM3":
                                self.optical_ports[3] = value
                            elif key == "OPTICAL_COM4":
                                self.optical_ports[4] = value
                            elif key == "DAC_COM1":
                                self.dec_ports[1] = value
                                self.dec_ports[2] = value
                            elif key == "DAC_COM2":
                                self.dec_ports[3] = value
                                self.dec_ports[4] = value


                logging.info(f"Loaded scanner ports: {self.scanner_ports}")
                logging.info(f"Loaded optical ports: {self.optical_ports}")
            except Exception as e:
                logging.error(f"Error loading COM port configuration: {e}")
                messagebox.showerror("Config Error", f"Failed to load COM port configuration: {e}")
        else:
            logging.warning(f"COM port config file not found at: {path}")
            messagebox.showwarning("Config Missing", 
                                 f"Configuration file not found: {path}\n"
                                 "Please create config.txt with SCANNER_COMx and OPTICAL_COMx mappings")

    def auto_scan_all_serials(self):
        """Auto scan serial numbers from all configured COM ports with enhanced logging"""
        self.status_label.config(text="Auto-scanning serial numbers...")
        
        # Update status indicators to show scanning
        for station_num in self.station_numbers.keys():
            if station_num in self.scanner_ports:
                ModernUI.update_status_indicator(self.status_indicators[station_num], "scanning")
        
        self.root.update()
        
        # Perform the scan
        scan_results = self.serial_scanner.scan_all_configured_ports()
        
        # Process scan results with enhanced logging
        successful_scans = 0
        failed_scans = 0
        retried_scans = 0
        for station_num, result in scan_results.items():
            if result['status'] == 'success' and result['serial']:
                self.station_numbers[station_num].set(result['serial'])
                self.finalize_serial_entry(station_num)
                successful_scans += 1
                
                # Log to Excel
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, result['serial'], "SERIAL_SCAN",
                    datetime.now(), datetime.now(), f"Scan {result['port']}", 
                    result['serial'], "PASS"
                )
                
                if result.get('retry_attempt', False):
                    retried_scans += 1
                    logging.info(f"Station {station_num}: Retry successful - Scanned {result['serial']} from {result['port']}")
                else:
                    logging.info(f"Station {station_num}: Scanned {result['serial']} from {result['port']}")
            else:
                failed_scans += 1
                self.failed_stations.add(station_num)
                error_msg = result.get('error', 'No data received')
                retry_text = " (after retry)" if result.get('retry_attempt', False) else ""
                
                # Log failure to Excel
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, "UNKNOWN", "SERIAL_SCAN",
                    datetime.now(), datetime.now(), f"Scan {result['port']}", 
                    "", "", "FAIL", error_msg
                )
                
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_action(f"Station {station_num} scan failed{retry_text}: {error_msg}", timestamp)
                
                logging.warning(f"Station {station_num}: Scan failed{retry_text} - {error_msg}")
                ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                
                for task_name in self.task_indicators[station_num]:
                    self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
        
        # Update status with detailed information
        status_parts = []
        if successful_scans > 0:
            status_parts.append(f"{successful_scans} successful")
        if retried_scans > 0:
            status_parts.append(f"{retried_scans} recovered on retry")
        if failed_scans > 0:
            status_parts.append(f"{failed_scans} failed")
        
        if status_parts:
            status_msg = f"Auto-scan completed: {', '.join(status_parts)}"
        else:
            status_msg = "Auto-scan completed: No serial numbers found"
        
        self.status_label.config(text=status_msg)
            
        logging.info(f"Scan summary: {successful_scans} successful ({retried_scans} after retry), {failed_scans} failed")
        if failed_scans > 0:
            logging.warning(f"Failed stations (after retry): {list(self.failed_stations)}")
        if retried_scans > 0:
            logging.info(f"Stations recovered on retry: {retried_scans}")

    def finalize_serial_entry(self, station_num):
        """Finalize a serial entry (lock the field and update status)"""
        entry_widget = self.entries[station_num]
        serial_value = self.station_numbers[station_num].get().strip()
        
        if serial_value:
            entry_widget.config(state="disabled", disabledbackground="#F5F5F5", disabledforeground="#333")
            self.scanned_serials.add(serial_value)
            ModernUI.update_status_indicator(self.status_indicators[station_num], "pending")

    def handle_manual_entry(self, event, station_num):
        """Handle manual serial number entry"""
        serial_value = self.station_numbers[station_num].get().strip()
        if serial_value:
            self.failed_stations.discard(station_num)
            self.finalize_serial_entry(station_num)

    def exit_fullscreen(self, event=None):
        self.root.attributes('-fullscreen', False)
        
    def create_header(self):
        # Create a frame for the header
        header_frame = tk.Frame(self.root, bg="#1976D2", height=72)
        header_frame.pack(fill=tk.X)
        
        # Logo frame on the left
        logo_frame = tk.Frame(header_frame, bg="#1976D2", width=200)
        logo_frame.pack(side=tk.LEFT, padx=20, pady=10)
        
        # Try to load and display the logo
        if os.path.exists("logo.png"):
            try:
                original_image = Image.open("logo.png")
                resized_image = original_image.resize((125, 40))
                logo_image = ImageTk.PhotoImage(resized_image)
                logo_label = tk.Label(logo_frame, image=logo_image, bg="#1976D2")
                logo_label.image = logo_image
            except:
                logo_label = tk.Label(logo_frame, text="POLARIS", font=("Arial", 20, "bold"), 
                                    bg="#1976D2", fg="white")
        else:
            logo_label = tk.Label(logo_frame, text="POLARIS", font=("Arial", 20, "bold"), 
                                bg="#1976D2", fg="white")

        logo_label.pack(side=tk.LEFT)
        
        # Add title
        title_label = tk.Label(header_frame, text="Smart Manufacturing Suite", 
                              font=("Arial", 18, "bold"), bg="#1976D2", fg="white")
        title_label.pack(side=tk.LEFT, padx=5, pady=15)

        toggle_frame = tk.Frame(header_frame, bg="#1976D2")
        toggle_frame.pack(side=tk.RIGHT, padx=20, pady=15)

        # Label before toggle button
        toggle_ng_programm_incomplate_label = tk.Label(toggle_frame, text="Pac File", 
                                font=("Arial", 18, "bold"), bg="#1976D2", fg="white")
        toggle_ng_programm_incomplate_label.pack(side=tk.LEFT, padx=(0, 10))  # small space before toggle

        # Toggle switch
        self.toggle = ToggleSwitch(toggle_frame, width=60, height=30)
        self.toggle.pack(side=tk.LEFT)

        toggle_ng_programm_complate_label = tk.Label(toggle_frame, text="Downloaded", 
                              font=("Arial", 18, "bold"), bg="#1976D2", fg="white")
        toggle_ng_programm_complate_label.pack(side=tk.RIGHT, padx=5, pady=15)

        
        # Add version info
        version_label = tk.Label(header_frame, text="V1.0.0.3 | 20-JAN-2026", 
                                font=("Arial", 10), bg="#1976D2", fg="#E1F5FE")
        version_label.pack(side=tk.LEFT, padx=5, pady=15)

        
        # Add toolbar
        toolbar_frame = tk.Frame(self.root, bg="#2196F3", height=55)
        toolbar_frame.pack(fill=tk.X)
        
        # Add buttons to toolbar
        run_btn_frame = ModernUI.create_modern_button(toolbar_frame, "▶ Run Test", self.run_test, 
                                                width=10, height=1, bg="#4CAF50", hover_bg="#388E3C")
        run_btn_frame.pack(side=tk.LEFT, padx=15, pady=7)

        
        # Bind Enter key to run_test function
        self.root.bind('<Return>', lambda event: self.run_test())

        stop_btn_frame = ModernUI.create_modern_button(toolbar_frame, "■ Stop", self.stop_test, 
                                                width=10, height=1, bg="#F44336", hover_bg="#D32F2F")
        stop_btn_frame.pack(side=tk.LEFT, padx=15, pady=7)

        new_batch_btn_frame = ModernUI.create_modern_button(toolbar_frame, "New Batch", self.show_start_batch_dialog, 
                                                        width=10, height=1, bg="#555259", hover_bg="#3b393d")
        new_batch_btn_frame.pack(side=tk.LEFT, padx=15, pady=7)
        self.toggle.disable()
        
    def export_excel_data(self):
        """Open the Excel files for viewing"""
        try:
            import subprocess
            import platform
            
            if platform.system() == 'Windows':
                subprocess.Popen(['start', 'excel', file_name], shell=True)
                subprocess.Popen(['start', 'excel', detailed_log_file], shell=True)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.Popen(['open', file_name])
                subprocess.Popen(['open', detailed_log_file])
            else:  # Linux
                subprocess.Popen(['xdg-open', file_name])
                subprocess.Popen(['xdg-open', detailed_log_file])
                
            messagebox.showinfo("Excel Export", f"Excel files opened:\n\n1. Main Results: {file_name}\n2. Detailed Logs: {detailed_log_file}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to open Excel files: {e}")

    def get_firmware_version_from_file(self):
        """Reads the current firmware version from the config file"""
        try:
            firmware_config = os.path.join("ConfigFiles", "firmware_config.txt")
            if os.path.exists(firmware_config):
                with open(firmware_config, 'r') as f:
                    version = f.read().strip()
                    return version
        except Exception as e:
            logging.error(f"Failed to read firmware version: {e}")
        return "Unknown"
        
    def create_main_content(self):
        # Create a frame for the main content with some padding
        main_frame = tk.Frame(self.root, bg="#F5F5F5")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create a frame for the test stations
        content_frame = tk.Frame(main_frame, bg="#F5F5F5")
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Dictionary to store serial numbers for each test station
        self.station_numbers = {}
        self.task_indicators = {}

        for i in range(4):
            station_num = i + 1

            # Create a frame for each test station
            station_frame = tk.Frame(content_frame, bg="white", bd=0)
            # station_frame.grid(row=row, column=col, padx=10, pady=10, sticky="nsew")
            station_frame.grid(row=0, column=i, padx=10, pady=10, sticky="nsew")
            
            # Add a canvas for rounded corners and shadow effect
            station_canvas = tk.Canvas(station_frame, bg="white", highlightthickness=0)
            station_canvas.pack(fill="both", expand=True)
            
            # Draw rounded rectangle for the station
            ModernUI.create_rounded_rectangle(station_canvas, 2, 2, 720, 1600, radius=10,
                                            fill="white", outline="#E0E0E0", width=2)
            
            # Create inner frame for content
            inner_frame = tk.Frame(station_canvas, bg="white")
            station_canvas.create_window(320, 400, window=inner_frame, width=620, height=800)
            
            # Station header
            header_frame = tk.Frame(inner_frame, bg="#2196F3", height=30)
            header_frame.pack(fill=tk.X)
            
            station_label = tk.Label(header_frame, text=f"Test Station {station_num}", 
                                    font=("Arial", 12, "bold"), bg="#2196F3", fg="white")
            station_label.pack(side=tk.LEFT, padx=15, pady=8)
            
            # Show configured COM port (optical port for testing)
            if station_num in self.optical_ports:
                port_label = tk.Label(header_frame, text=f"({self.optical_ports[station_num]})", 
                                    font=("Arial", 10), bg="#2196F3", fg="#E1F5FE")
                port_label.pack(side=tk.RIGHT, padx=15, pady=8)
            
            # Create status indicator
            status_indicator = ModernUI.create_status_indicator(station_canvas, size=50, initial_state="pending")
            status_indicator.place(x=250, y=150)
            self.status_indicators[station_num] = status_indicator
            
            # Serial number frame
            serial_frame = tk.Frame(inner_frame, bg="white", height=50)
            serial_frame.pack(fill=tk.X, padx=10, pady=(15, 5))
            
            serial_label = tk.Label(serial_frame, text="Serial Number:", bg="white", font=("Arial", 11))
            serial_label.pack(side=tk.LEFT, padx=(5, 5))
            
            # Entry field for serial number
            self.station_numbers[station_num] = tk.StringVar()
            entry = tk.Entry(serial_frame, textvariable=self.station_numbers[station_num], 
                           width=28, font=("Arial", 11), bd=2, relief=tk.GROOVE)
            entry.pack(side=tk.LEFT, padx=2)
            entry.bind("<Return>", lambda event, s=station_num: self.handle_manual_entry(event, s))
            self.entries[station_num] = entry
            
            # Task status frame
            task_frame = tk.Frame(inner_frame, bg="white")
            task_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
            
            # List of tasks
            tasks = [
                "Mapped with COM Port",
                "BOOT Mode Initialization",
                "BOOT MODE ON",
                "Buck Output",
                "LDO Output",
                "1V8",
                "IMG SINK Programming",
                "N58 Programming Initialization",
                "Super Cap Output",
                "BOOT MODE OFF",
                "MODULE SUPPLY ON",
                "Firmware Version Check",
                "MODULE SUPPLY OFF",
                "SUPERCAP DISCHARGE",
                "SUPERCAP DISCHARGED VOLTAGE",
                "Module Supply Disconnected"
            ]
            
            # Store task indicators for this station
            self.task_indicators[station_num] = {}
            
            # Display tasks with status indicators
            for task in tasks:
                task_indicator, icon_label = ModernUI.create_task_indicator(task_frame, task, initial_state="pending")
                task_indicator.pack(fill=tk.X, pady=1)
                self.task_indicators[station_num][task] = icon_label
        
        for col in range(4):
            content_frame.columnconfigure(col, weight=1)
        content_frame.rowconfigure(0, weight=1)
    
    def create_status_bar(self):
        status_frame = tk.Frame(self.root, bg="#E0E0E0", height=30)
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)

        self.status_label = tk.Label(status_frame, text="Ready", anchor="w", bg="#E0E0E0", font=("Arial", 10))
        self.status_label.pack(side=tk.LEFT, padx=15, pady=5)

        self.batch_time = tk.Label(status_frame, text="BATCH TIME: 00:00", bg="#4CAF50", fg="white", 
                                padx=10, pady=5, font=("Arial", 10, "bold"))
        self.batch_time.pack(side=tk.RIGHT, padx=15)

        self.status = tk.Label(status_frame, text="OFFLINE", bg="#FFD700", fg="#333", 
                            padx=10, pady=5, font=("Arial", 10, "bold"))
        self.status.pack(side=tk.RIGHT, padx=15)

        # Create labels to display counters in the UI
        # self.cycle_count_label = tk.Label(status_frame, text="Total Cycles: 0", bg="#E0E0E0", 
        #                                 fg="#333", font=("Arial", 10, "bold"))
        # self.cycle_count_label.pack(side=tk.RIGHT, padx=15)

        # self.passed_stations_label = tk.Label(status_frame, text="Passed PCBA: 0", bg="#E0E0E0", 
        #                                     fg="#333", font=("Arial", 10, "bold"))
        # self.passed_stations_label.pack(side=tk.RIGHT, padx=15)

        
        # Firmware version display
        self.firmware_version = self.get_firmware_version_from_file()
        self.firmware_version_label = tk.Label(status_frame, text=f"Current Firmware version: {self.firmware_version}", 
                                            bg="#E0E0E0", fg="#333", font=("Arial", 10 , "bold"))
        self.firmware_version_label.pack(side=tk.RIGHT, padx=10)

    def gather_com_ports_serial_numbers(self):
        """Map serial numbers to COM ports with enhanced logging"""
        self.status_label.config(text="Mapping COM ports to serial numbers...")
        self.root.update()

        self.active_ports_serials = []
        self.toggle.turn_off()

        for station_num in range(1, 5):
            if station_num in self.failed_stations:
                continue

            serial_number = self.station_numbers[station_num].get().strip()
            if not serial_number:
                self.task_indicators[station_num]["Mapped with COM Port"].config(
                    text="⛔ Skipped", fg="#9E9E9E")
                self.failed_stations.add(station_num)
                continue

            optical_com_port = self.optical_ports.get(station_num)



            if optical_com_port:
                self.active_ports_serials.append((optical_com_port, serial_number, station_num))


                # Log COM port mapping to Excel
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, serial_number, "COM_PORT_MAPPING",
                    datetime.now(), datetime.now(), f"Map to {optical_com_port}", 
                    optical_com_port, "PASS"
                )
                if self.eng_var.get():
                    self.station_numbers[station_num].set(f"{optical_com_port} - {serial_number}")
                    self.entries[station_num].config(
                        state="disabled", disabledbackground="#F5F5F5", disabledforeground="#333")

                self.task_indicators[station_num]["Mapped with COM Port"].config(
                    text="✓", fg="#4CAF50")
                self.station_test_results[serial_number] = {}
                self.station_test_results[serial_number]["COM_PORT_MAPPING"] = {"status": "PASS", "value": ""}
            else:
                # Log mapping failure
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, serial_number, "COM_PORT_MAPPING",
                    datetime.now(), datetime.now(), "No optical port configured", 
                    "", "", "FAIL", f"No optical port configured for station {station_num}"
                )
                
                self.task_indicators[station_num]["Mapped with COM Port"].config(
                    text="✕", fg="#F44336")
                self.failed_stations.add(station_num)

                self.station_test_results[serial_number]["COM_PORT_MAPPING"] = {"status": "FAIL", "value": ""}

        logging.warning(f"Failed stations: {list(self.failed_stations)}")
        logging.info(f"Mapped COM Ports and Serial Numbers for optical tests: {self.active_ports_serials}")
        logging.info(f"Skipped failed stations: {list(self.failed_stations)}")
        self.status_label.config(text="COM port mapping complete")

    def check_firmware_log(self, output):
        """Extract verification results from firmware flash output"""
        pattern = r"#(\d+):\s+\.\.\s+(OK|failed)"
        matches = re.findall(pattern, output)
        ret_dict = {f"#{num}": status for num, status in matches}
        l = ['#1','#2','#5','#4']
        for num in l:
            if num not in ret_dict.keys():
                ret_dict[num] = 'failed'
        return ret_dict

    def FirmwareFlash(self, is_main_firmware=False):
        """Enhanced firmware flash with detailed logging and station isolation"""
        task_name = "Main Firmware Download" if is_main_firmware else "IMG SINK Programming"
        Flashfile = "MainFirmwareFlash.py" if is_main_firmware else "FirmwareFlash.py"
        self.status_label.config(text=f"Running {task_name.lower()}...")

        with self.test_completion_lock:
            self.completed_tests["IMG SINK Programming"] = False

        active_stations_for_flash = []
        for station_num in range(1, 5):
            if station_num in self.failed_stations:
                continue
            serial_number = self.station_numbers[station_num].get().strip()
            if serial_number:
                active_stations_for_flash.append((station_num, serial_number))
                self.task_indicators[station_num][task_name].config(text="⏳", fg="#2196F3")

        self.root.update()

        folder_path = r'ConfigFiles'
        Config_file = 'img_lastused_gparm.cfg'
        os.path.join(os.path.dirname(os.path.abspath(__file__)), Flashfile)
        script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), Flashfile)

        cmd = [sys.executable, script_path, os.path.join(folder_path, Config_file)]

        start_time = datetime.now()
        logging.info(f"Executing Firmware Flash command: {' '.join(cmd)}")

        flash_success = False
        # raw_output = ""
        raw_output = "D O N E"
        error_output = ""
        result_station_dict = {}

        try:
            process = 0
            process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                encoding='utf-8',
                text=True,
                cwd=os.path.dirname(script_path)
            )

            raw_output, error_output = process.communicate()
            logging.debug(f"Firmware Flash raw output:\n{raw_output}")
            logging.debug(f"Firmware Flash error output:\n{error_output}")
            # if process == 0:
            #     if "D O N E" in raw_output:
            #         flash_success = True
            #         logging.info("Firmware flash completed successfully.")
            #     else:
            #         logging.error("Firmware flash exited with success code but 'D O N E' not found.")
            # else:
            #     logging.error(f"Firmware flash failed with return code {process}")

            if process.returncode == 0 and not error_output:
                if "D O N E" in raw_output:
                    result_station_dict = self.check_firmware_log(raw_output)
                    logging.info(result_station_dict)
                    flash_success = True
                    logging.info("Firmware flash completed successfully.")
                else:
                    logging.error("Firmware flash exited with success code but 'D O N E' not found.")
            else:
                logging.error(f"Firmware flash failed with return code {process.returncode}: {error_output}")

        except Exception as e:
            error_output = str(e)
            logging.error(f"Failed to execute firmware flash: {e}")
        finally:
            end_time = datetime.now()

            for station_num, serial_number in active_stations_for_flash:
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, serial_number, task_name,
                    start_time, end_time, ' '.join(cmd), raw_output[:200],
                    "OK" if flash_success else "FAIL", error_output
                )

                with self.test_completion_lock:
                    if serial_number not in self.station_test_results:
                        self.station_test_results[serial_number] = {}
                    if serial_number not in self.station_test_times:
                        self.station_test_times[serial_number] = {'start': start_time, 'end': end_time}

                    if flash_success:
                        if (str(station_num) == "1") and (result_station_dict.get("#1") == "OK"):
                            self.station_test_results[serial_number][task_name] = {"status": "OK", "value": "N/A"}
                            self.task_indicators[station_num][task_name].config(text="✓", fg="#4CAF50")
                            if station_num in self.failed_stations:
                                self.failed_stations.remove(station_num)
                                self.station_stop_flags[station_num] = False
                                self.is_break = False
                        elif (str(station_num) == "2") and (result_station_dict.get("#2") == "OK"):
                            self.station_test_results[serial_number][task_name] = {"status": "OK", "value": "N/A"}
                            self.task_indicators[station_num][task_name].config(text="✓", fg="#4CAF50")
                            if station_num in self.failed_stations:
                                self.failed_stations.remove(station_num)
                                self.station_stop_flags[station_num] = False
                                self.is_break = False
                        elif (str(station_num) == "3") and (result_station_dict.get("#4") == "OK"):
                            self.station_test_results[serial_number][task_name] = {"status": "OK", "value": "N/A"}
                            self.task_indicators[station_num][task_name].config(text="✓", fg="#4CAF50")
                            if station_num in self.failed_stations:
                                self.failed_stations.remove(station_num)
                                self.station_stop_flags[station_num] = False
                                self.is_break = False
                        elif (str(station_num) == "4") and (result_station_dict.get("#5") == "OK"):
                            self.station_test_results[serial_number][task_name] = {"status": "OK", "value": "N/A"}
                            self.task_indicators[station_num][task_name].config(text="✓", fg="#4CAF50")
                            if station_num in self.failed_stations:
                                self.failed_stations.remove(station_num)
                                self.station_stop_flags[station_num] = False
                                self.is_break = False
                        else:
                            self.station_test_results[serial_number][task_name] = {"status": "FAIL", "value": "N/A"}
                            self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
                            if station_num in self.failed_stations:
                                self.failed_stations.add(station_num)
                                self.station_stop_flags[station_num] = True
                                self.is_break = False
                    else:
                        self.station_test_results[serial_number][task_name] = {"status": "FAIL", "value": "N/A"}
                        self.task_indicators[station_num][task_name].config(text="✕", fg="#F44336")
                        # Mark this station as failed in programming phase
                        self.programming_phase_failed = True
                        self.failed_stations.add(station_num)
                        self.station_stop_flags[station_num] = True
                        self.is_break = True

            with self.test_completion_lock:
                self.completed_tests["Test Firmware Download"] = True

            if flash_success:
                self.status_label.config(text="✅ Firmware Flash Successful. Proceeding with tests.")
                # REMOVED POWER RESET DIALOG - Replace with 14-second sleep
                self.status_label.config(text="Waiting 4 seconds for device reset...")
                self.root.update()
                time.sleep(4)  # 14-second sleep instead of power reset dialog
                return True

            # If firmware flash failed, force fail & stop testing for affected stations
            if not flash_success:
                logging.warning("Firmware flash failed — stopping all further tests for affected stations.")
                self.status_label.config(text="❌ Firmware Flash Failed. Affected stations will be skipped.")
                self.status_label.update_idletasks()
                return False

    def get_provisioning_parameters(self, data):
        # Extract block between summary header and underline line
        block_pattern = r"Provisioned NIC Parameters Summary:\s*_{80,}\s*(.*?)_{80,}"
        block_match = re.search(block_pattern, data, re.DOTALL)
        
        if not block_match:
            return {}

        extracted_block = block_match.group(1)

        # Extract key-value pairs
        kv_pattern = r"'([\w_]+)'\s*:\s*(?:'([^']*)'|\[None\]|(\d+))"
        matches = re.findall(kv_pattern, extracted_block)

        result = {}
        for key, str_val, num_val in matches:
            if num_val:
                result[key] = int(num_val)
            elif str_val == "Not set":
                result[key] = None
            elif str_val:
                result[key] = str_val.strip()
            else:
                result[key] = None

        return result

    def run_test_provisioning(self, cmd, station_num, test_type, task_name):
        cmd_test = cmd + ["--test_type", test_type]
        timeout_duration = 20
        print(cmd_test)
        stdout, stderr, return_code, timed_out = self.process_manager.run_process_with_timeout(
            cmd_test, timeout_duration, station_num, task_name
        )
        raw_output = stdout + stderr
        return {
            "test_type": test_type,
            "stdout": stdout,
            "stderr": stderr,
            "return_code": return_code,
            "timed_out": timed_out,
            "raw_output": raw_output
        }

    def _run_test_script_parallel(self, script_name, task_name, result_key, com_port = None, serial_number = None, station_num = None, value_key=None):
        """Enhanced test script runner with improved value parsing and station isolation"""
        print(task_name,'----------------')
        is_firmware_script = False
        is_execute = False
        # Check if this station has been marked as failed
        # if station_num in self.failed_stations:
        #     logging.info(f"Station {station_num} already failed, skipping {task_name}")
        #     return serial_number, "FAIL", station_num, task_name

        # # Check if we should stop testing for this station
        # if self.station_stop_flags.get(station_num, False):
        #     logging.info(f"Stop flag set for Station {station_num}, skipping {task_name}")
        #     return serial_number, "FAIL", station_num, task_name
        if "BOOT Mode Initialization" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "01"]
            is_firmware_script = False
            is_execute = True
        elif "BOOT MODE ON" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "05"]
            is_firmware_script = False
            is_execute = True
        elif "IMG SINK Programming" in task_name:
            cmd = []
            is_firmware_script = True
            is_execute = True
        elif "Firmware Version Check" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", "IMG_Provisioning_Testing_Scripts", "read_img_parameters.py")
            cmd = [sys.executable, script_path, "--serial_port", com_port, "--baudrate", str(self.baudrate)]
            is_firmware_script = False
            is_execute = True
        elif "N58 Programming Initialization" in task_name:
            cmd = []
            is_firmware_script = False
            is_execute = False
        elif "BOOT MODE OFF" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "00"]
            is_firmware_script = False
            is_execute = True
        elif "MODULE SUPPLY ON" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "04"]
            is_firmware_script = False
            is_execute = True
        elif "MODULE SUPPLY OFF" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "0"]
            is_firmware_script = False
            is_execute = True
        elif "SUPERCAP DISCHARGE" == task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "2"]
            is_firmware_script = False
            is_execute = True

        elif "Super Cap Output" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--deviceDescription", "USB-4704,BID#"]
            is_firmware_script = False
            is_execute = True
            time.sleep(2)  # Short delay to ensure device is ready

        elif "Buck Output" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--deviceDescription", "USB-4704,BID#"]
            is_firmware_script = False
            is_execute = True
            time.sleep(2)  # Short delay to ensure device is ready
        elif "SUPERCAP DISCHARGED VOLTAGE" == task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--deviceDescription", "USB-4704,BID#"]
            is_firmware_script = False
            is_execute = True
            time.sleep(2)  # Short delay to ensure device is ready
        
        elif "Module Supply Disconnected" in task_name:
            script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", script_name)
            cmd = [sys.executable, script_path, "--inputVal", "0"]
            is_firmware_script = False
            is_execute = True
        

        else:
            if task_name == 'GPIO Test':
                script_name = "testing.py"
                script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", "IMG_Provisioning_Testing_Scripts", script_name)
                cmd = [sys.executable, script_path, "--serial_port", com_port,"--test_router_address", "3", "--baudrate" ,"19200","--network_address", "0xF0EE0F","--network_channel", "3","--encryption_key", "33333333333333333333333333333333","--authentication_key", "44444444444444444444444444444444"]
            else:
                script_name = "provisioning.py"
                script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts", "tests", "IMG_Provisioning_Testing_Scripts", script_name)
                cmd = [sys.executable, script_path, "--serial_number", serial_number, "--serial_port", "COM3"]
            is_firmware_script = False
            is_execute = True
        
        
        start_time = datetime.now()
        logging.info(f"Station {station_num} ({serial_number}): Executing {task_name} command: {' '.join(cmd)}")
        
        test_status = "FAIL"
        test_value = "N/A"
        raw_output = ""
        error_message = ""
        # sim1_cellular_test_status = "FAIL"
        # sim2_cellular_test_status = "FAIL"
        gpio_test_status = "FAIL"
        external_flash_test_status = "FAIL"
        rx_rssi_value = ""
        tx_rssi_value = ""
        rf_test_status = "FAIL"
        buck_voltage = 0
        ldo_output = 0
        super_cap_voltage = 0
        v8_voltage = 0
        data_dict = {}
        core_ver = None
        main_ver = None
        sink_app_ver = None
        modem_firmware_ver = None
        meter_lib_ver = None
        core_ver_status = None
        main_ver_status = None
        sink_app_ver_status = None
        modem_firmware_ver_status = None
        meter_lib_ver_status = None

        
        logging.info(f"[RUN] Station {station_num} | Task: {task_name} | Port: {com_port} | Serial: {serial_number}")

        try:
            logging.debug(f"[SPAWNED] Started {task_name} for Station {station_num} on {com_port}")

            if is_firmware_script:
                if self.FirmwareFlash(is_main_firmware=False):
                    test_status = "PASS"
                else:
                    test_status  = "FAIL"
            
            else:
                # Enhanced timeout handling - increased timeout for All Test
                timeout_duration = 30
                
                # Use ProcessManager for robust process handling
                if is_execute:
                    if task_name == "GPIO Test":

                        cmd_gpio = cmd + ["--test_type", "gpio"]
                        stdout_gpio, stderr_gpio, return_code_gpio, timed_out_gpio = self.process_manager.run_process_with_timeout(
                            cmd_gpio, timeout_duration, station_num, task_name
                        )
                        raw_output_gpio = stdout_gpio + stderr_gpio

                        cmd_external_flash = cmd + ["--test_type", "external_flash"]
                        stdout_external_flash, stderr_external_flash, return_code_external_flash, timed_out_external_flash = self.process_manager.run_process_with_timeout(
                            cmd_external_flash, timeout_duration, station_num, task_name
                        )
                        raw_output_external_flash = stdout_external_flash + stderr_external_flash

                        cmd_rf_rssi = cmd + ["--test_type", "rf_rssi"]
                        stdout_rf_rssi, stderr_rf_rssi, return_code_rf_rssi, timed_out_rf_rssi = self.process_manager.run_process_with_timeout(
                            cmd_rf_rssi, timeout_duration, station_num, task_name
                        )
                        raw_output_rf_rssi = stdout_rf_rssi + stderr_rf_rssi

                        timed_out = timed_out_rf_rssi
                    else:
                        stdout, stderr, return_code, timed_out = self.process_manager.run_process_with_timeout(
                            cmd, timeout_duration, station_num, task_name
                        )
                        raw_output = stdout + stderr

                    
                    # print(raw_output)
                    logging.debug(f"{task_name} raw output for Station {station_num}:\n{raw_output}")

                    if timed_out:
                        error_message = f"Test {task_name} timed out and was forcefully terminated after {timeout_duration} seconds"
                        logging.error(f"⏰ Station {station_num}: {error_message}")
                        test_status = "FAIL"
                        test_value = "Timeout - Process Killed"
                        
                        # Mark this station as failed for timeout
                        self.failed_stations.add(station_num)
                        self.station_stop_flags[station_num] = True
                        
                        return serial_number, test_status, station_num, task_name
                else:
                    test_status = "PASS"

                # Enhanced parsing based on your log data
                if ("BOOT Mode Initialization" in task_name) or ("BOOT MODE" in task_name) or ("SUPERCAP DISCHARGE" == task_name) or ("MODULE SUPPLY ON" in task_name) or ("MODULE SUPPLY OFF" in task_name) or ("Module Supply Disconnected" in task_name):
                    status_match = re.search(r"DO output completed!", raw_output)
                    if status_match:
                        # test_value = int(version_match.group(1))
                        test_status = "PASS"
                    else:
                        # Fallback to parsed version from logs
                        test_status = "FAIL"
                            
                elif ("Buck Output" in task_name) or ("SUPERCAP DISCHARGED VOLTAGE" in task_name) or ("Super Cap Output" in task_name):
                    # Look for "NIC Status      : 1" or "NIC Status      : 0"
                    lines = [line.strip() for line in raw_output.splitlines() if line.strip()]

                    channel_data = {}
                    data_dict = ast.literal_eval(lines[0])

                    if len(data_dict.keys()) >= 16:
                        test_status = "PASS"
                        
                    else:
                        test_status = "FAIL"
                elif "Firmware Version Check" in task_name:
                    # print(raw_output,'------------------------------------- com_port',com_port)
                    read_parm_dict = self.get_provisioning_parameters(raw_output)
                    

                    core_ver = read_parm_dict['GW_CORE_VER']
                    main_ver = read_parm_dict['GW_MAIN_VER']
                    sink_app_ver = read_parm_dict['GW_SINK_APP_VER']
                    modem_firmware_ver = read_parm_dict['GW_MODEM_FIRMWARE_VER']
                    meter_lib_ver = read_parm_dict['GW_METER_LIB_VER']
                    if core_ver == self.core_ver:
                        core_ver_status = "PASS"
                    else:
                        core_ver_status = "FAIL"

                    if main_ver == self.main_ver:
                        main_ver_status = "PASS"
                    else:
                        main_ver_status = "FAIL"

                    if sink_app_ver == self.sink_app_ver:
                        sink_app_ver_status = "PASS"
                    else:
                        sink_app_ver_status = "FAIL"

                    if modem_firmware_ver == self.modem_firmware_ver:
                        modem_firmware_ver_status = "PASS"
                    else:
                        modem_firmware_ver_status = "FAIL"

                    if meter_lib_ver == self.meter_lib_ver:
                        meter_lib_ver_status = "PASS"
                    else:
                        meter_lib_ver_status = "FAIL"
                    
                    if all([
                        core_ver_status == "PASS",
                        main_ver_status == "PASS",
                        sink_app_ver_status == "PASS",
                        modem_firmware_ver_status == "PASS",
                        meter_lib_ver_status == "PASS"
                    ]):
                        test_status = "PASS"
                        test_value = ""
                    else:
                        test_status = "FAIL"
                        test_value = ""

                
                elif "GPIO Test" in task_name:
                    test_status = "FAIL"
                    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')

                    clean_output_gpio = ansi_escape.sub('', raw_output_gpio)
                    clean_output_external_flash = ansi_escape.sub('', raw_output_external_flash)
                    clean_output_rf_rssi = ansi_escape.sub('', raw_output_rf_rssi)

                    # Save result to file
                    result_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Detailed Test Result")
                    os.makedirs(result_folder, exist_ok=True)
                    
                    log_folder = os.path.join(result_folder, f"{current_date}")
                    os.makedirs(log_folder, exist_ok=True)
                    
                    result_file_name = f"{serial_number}_SIM_Test_result.txt"
                    result_file_path = os.path.join(log_folder, result_file_name)
                    mode = 'a' if os.path.exists(result_file_path) else 'w'
                    with open(result_file_path, mode, encoding="utf-8") as result_file:
                        if mode == 'a':
                            result_file.write("\n--- Updated at {} ---\n".format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

                        result_file.write(clean_output_gpio)
                        result_file.write(clean_output_external_flash)
                        result_file.write(clean_output_rf_rssi)
                        
                    # Parsing logic
                    print(clean_output_gpio,'---------------->', station_num)
                    if "GPIO Test:" in clean_output_gpio and "Success" in clean_output_gpio.split("GPIO Test:")[1].splitlines()[0]:
                        gpio_test_status = "PASS"
                    if "External Flash Test:" in clean_output_external_flash and "Success" in clean_output_external_flash.split("External Flash Test:")[1].splitlines()[0]:
                        external_flash_test_status = "PASS"
                    if "RF Test:" in clean_output_rf_rssi and "Success" in clean_output_rf_rssi.split("RF Test:")[1].splitlines()[0]:
                        rf_test_status = "PASS"
                    print(gpio_test_status,'---------------->', station_num)
                    
                    # print(clean_output)
                    # Extract RSSI values using regex
                    rx_match = re.search(r"RX RSSI:\s*([-\d]+ dBm)", clean_output_rf_rssi)
                    tx_match = re.search(r"TX RSSI:\s*([-\d]+ dBm)", clean_output_rf_rssi)

                    if rx_match:
                        rx_rssi_value = rx_match.group(1)

                    if tx_match:
                        tx_rssi_value = tx_match.group(1)


                    # if "Overall Test Result: ❌ FAIL ❌" in clean_output:
                    if(gpio_test_status == "PASS") and (external_flash_test_status == "PASS") and (rf_test_status == "PASS"):
                        test_status = "PASS"
                    else:
                        test_status = "FAIL"
        
                else:
                    test_status = "PASS"
        except Exception as e:
            error_message = f"Test {task_name} failed with exception: {str(e)}"
            logging.error(f"Station {station_num}: {error_message}")
            test_status = "FAIL"
            test_value = "N/A"
            
            # Mark this station as failed for exceptions
            self.failed_stations.add(station_num)
            self.station_stop_flags[station_num] = True
            
            return serial_number, "FAIL", station_num, task_name
        
        end_time = datetime.now()

        with self.test_completion_lock:
            if task_name == "IMG SINK Programming":
                # For programming, we already logged inside FirmwareFlash
                return serial_number, test_status, station_num, task_name
            if serial_number is not None:
                if serial_number not in self.station_test_results:
                    self.station_test_results[serial_number] = {}
                if serial_number not in self.station_test_times:
                    self.station_test_times[serial_number] = {'start': start_time, 'end': end_time}
                else:
                    self.station_test_times[serial_number]['end'] = end_time
                self.station_test_results[serial_number][task_name] = {"status": test_status, "value": test_value}
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, serial_number, task_name,
                    start_time, end_time, ' '.join(cmd), raw_output[:200],
                    "PASS" if test_status else "FAIL", ""
                )
                if task_name == "Firmware Version Check":
                    self.station_test_results[serial_number]["GW_CORE_VER"] = {"status": core_ver_status, "value": core_ver}
                    self.station_test_results[serial_number]["GW_MAIN_VER"] = {"status": main_ver_status, "value": main_ver}
                    self.station_test_results[serial_number]["GW_SINK_APP_VER"] = {"status": sink_app_ver_status, "value": sink_app_ver}
                    self.station_test_results[serial_number]["GW_MODEM_FIRMWARE_VER"] = {"status": modem_firmware_ver_status, "value": modem_firmware_ver}
                    self.station_test_results[serial_number]["GW_METER_LIB_VER"] = {"status": meter_lib_ver_status, "value": meter_lib_ver}
            
                if task_name == "GPIO Test":
                #     self.station_test_results[serial_number]["sim1_cellular_test_status"] = {"status": sim1_cellular_test_status, "value": ""}
                #     self.station_test_results[serial_number]["sim2_cellular_test_status"] = {"status": sim2_cellular_test_status, "value": ""}
                    self.station_test_results[serial_number]["gpio_test_status"] = {"status": gpio_test_status, "value": ""}
                    self.station_test_results[serial_number]["external_flash_test_status"] = {"status": external_flash_test_status, "value": ""}
                    self.station_test_results[serial_number]["rx_rssi_value"] = {"status": rf_test_status, "value": rx_rssi_value}
                    self.station_test_results[serial_number]["tx_rssi_value"] = {"status": rf_test_status, "value": tx_rssi_value}
                    self.station_test_results[serial_number]["rf_test_status"] = {"status": rf_test_status, "value": ""}

                    # self.root.after(0, lambda s=station_num, t="SIM1 Cellular Test", stat=sim1_cellular_test_status: self._update_task_ui(station_num, "SIM1 Cellular Test", sim1_cellular_test_status))
                    # self.root.after(0, lambda s=station_num, t="SIM2 Cellular Test", stat=sim2_cellular_test_status: self._update_task_ui(station_num, "SIM2 Cellular Test", sim2_cellular_test_status))
                    self.root.after(0, lambda s=station_num, t="GPIO Test", stat=gpio_test_status: self._update_task_ui(station_num, "GPIO Test", gpio_test_status))
                    self.root.after(0, lambda s=station_num, t="External Flash Test", stat=external_flash_test_status: self._update_task_ui(station_num, "External Flash Test", external_flash_test_status))
                    self.root.after(0, lambda s=station_num, t="RF RSSI Test", stat=rf_test_status: self._update_task_ui(station_num, "RF RSSI Test", rf_test_status))
                    self.root.after(0, partial(self._update_task_ui, station_num, "GPIO Test", gpio_test_status))

            else:
                for com_port, serial_number, station_num in self.active_ports_serials:
                    if serial_number.strip():
                        if serial_number not in self.station_test_results:
                            self.station_test_results[serial_number] = {}
                        if serial_number not in self.station_test_times:
                            self.station_test_times[serial_number] = {'start': start_time, 'end': end_time}
                        else:
                            self.station_test_times[serial_number]['end'] = end_time
                        self.station_test_results[serial_number][task_name] = {"status": test_status, "value": test_value}
                        self.excel_logger.log_individual_test(
                            self.session_id, station_num, serial_number, task_name,
                            start_time, end_time, ' '.join(cmd), raw_output[:200],
                            "PASS" if test_status else "FAIL", ""
                        )
                        if task_name == "Firmware Version Check":
                            self.station_test_results[serial_number]["GW_CORE_VER"] = {"status": core_ver_status, "value": core_ver}
                            self.station_test_results[serial_number]["GW_MAIN_VER"] = {"status": main_ver_status, "value": main_ver}
                            self.station_test_results[serial_number]["GW_SINK_APP_VER"] = {"status": sink_app_ver_status, "value": sink_app_ver}
                            self.station_test_results[serial_number]["GW_MODEM_FIRMWARE_VER"] = {"status": modem_firmware_ver_status, "value": modem_firmware_ver}
                            self.station_test_results[serial_number]["GW_METER_LIB_VER"] = {"status": meter_lib_ver_status, "value": meter_lib_ver}
                    
                        if task_name == "Buck Output":
                            buck_voltage = float(data_dict["Channel " + str((station_num - 1)*4 + 0 ) + " data"])
                            ldo_output = float(data_dict["Channel " + str((station_num - 1)*4 + 1 ) + " data"])
                            super_cap_voltage = float(data_dict["Channel " + str((station_num - 1)*4 + 2 ) + " data"])
                            v8_voltage = float(data_dict["Channel " + str((station_num - 1)*4 + 3 ) + " data"])

                            if (float(buck_voltage) >= self.min_buck_voltage) and (float(buck_voltage) <= self.max_buck_voltage):
                                self.station_test_results[serial_number]["Buck Output"] = {"status": "PASS", "value": data_dict["Channel " + str((station_num - 1)*4 + 0 ) + " data"]}
                                self.root.after(0, lambda s=station_num, t="Buck Output", stat="PASS": self._update_task_ui(s, "Buck Output", "PASS"))
                            else:
                                self.station_test_results[serial_number]["Buck Output"] = {"status": "FAIL", "value": data_dict["Channel " + str((station_num - 1)*4 + 0 ) + " data"]}
                                self.root.after(0, lambda s=station_num, t="Buck Output", stat="FAIL": self._update_task_ui(s, "Buck Output", "FAIL"))
                            if (ldo_output >= self.min_ldo_output) and (ldo_output <= self.max_ldo_output):
                                self.station_test_results[serial_number]["LDO output"] = {"status": "PASS", "value": data_dict["Channel " + str((station_num - 1)*4 + 1) + " data"]}
                                self.root.after(0, lambda s=station_num, t="LDO Output", stat="PASS": self._update_task_ui(s, "LDO Output", "PASS"))
                            else:
                                self.station_test_results[serial_number]["LDO output"] = {"status": "FAIL", "value": data_dict["Channel " + str((station_num - 1)*4 + 1) + " data"]}
                                self.root.after(0, lambda s=station_num, t="LDO Output", stat="FAIL": self._update_task_ui(s, "LDO Output", "FAIL"))
                            if (v8_voltage >= self.min_v8_voltage) and (v8_voltage <= self.max_v8_voltage):
                                self.station_test_results[serial_number]["1V8"] = {"status": "PASS", "value": data_dict["Channel " + str((station_num - 1)*4 + 3) + " data"]}
                                self.root.after(0, lambda s=station_num, t="1V8", stat="PASS": self._update_task_ui(s, "1V8", "PASS"))
                            else:
                                self.station_test_results[serial_number]["1V8"] = {"status": "FAIL", "value": data_dict["Channel " + str((station_num - 1)*4 + 3) + " data"]}
                                self.root.after(0, lambda s=station_num, t="1V8", stat="FAIL": self._update_task_ui(s, "1V8", "FAIL"))
                        
                        if task_name == "Super Cap Output":
                            super_cap_voltage = float(data_dict["Channel " + str((station_num - 1)*4 + 2 ) + " data"])

                            if (super_cap_voltage >= self.min_super_cap_voltage) and (super_cap_voltage <= self.max_super_cap_voltage):
                                self.station_test_results[serial_number]["Super Cap Voltage"] = {"status": "PASS", "value": data_dict["Channel " + str((station_num - 1)*4 + 2) + " data"]}
                                self.root.after(0, lambda s=station_num, t="Super Cap Voltage", stat="PASS": self._update_task_ui(s, "Super Cap Voltage", "PASS"))
                            else:
                                self.station_test_results[serial_number]["Super Cap Voltage"] = {"status": "FAIL", "value": data_dict["Channel " + str((station_num - 1)*4 + 2) + " data"]}
                                self.root.after(0, lambda s=station_num, t="Super Cap Voltage", stat="FAIL": self._update_task_ui(s, "Super Cap Voltage", "FAIL"))
                        
                        if task_name == "SUPERCAP DISCHARGED VOLTAGE":
                            super_cap_discharged_voltage = float(data_dict["Channel " + str((station_num - 1)*4 + 2 ) + " data"])

                            if (float(super_cap_discharged_voltage) >= self.min_super_cap_discharge_voltage) and (float(super_cap_discharged_voltage) <= self.max_super_cap_discharge_voltage):
                                self.station_test_results[serial_number]["SUPERCAP DISCHARGED VOLTAGE"] = {"status": "PASS", "value": data_dict["Channel " + str((station_num - 1)*4 + 2 ) + " data"]}
                                self.root.after(0, lambda s=station_num, t="SUPERCAP DISCHARGED VOLTAGE", stat="PASS": self._update_task_ui(s, t, "PASS"))
                            else:
                                self.station_test_results[serial_number]["SUPERCAP DISCHARGED VOLTAGE"] = {"status": "FAIL", "value": data_dict["Channel " + str((station_num - 1)*4 + 2 ) + " data"]}
                                self.root.after(0, lambda s=station_num, t="SUPERCAP DISCHARGED VOLTAGE", stat="FAIL": self._update_task_ui(s, t, "FAIL"))

                        if task_name == "GPIO Test":
                            # self.station_test_results[serial_number]["sim1_cellular_test_status"] = {"status": sim1_cellular_test_status, "value": ""}
                            # self.station_test_results[serial_number]["sim2_cellular_test_status"] = {"status": sim2_cellular_test_status, "value": ""}
                            self.station_test_results[serial_number]["gpio_test_status"] = {"status": gpio_test_status, "value": ""}
                            self.station_test_results[serial_number]["external_flash_test_status"] = {"status": external_flash_test_status, "value": ""}
                            self.station_test_results[serial_number]["rx_rssi_value"] = {"status": rf_test_status, "value": rx_rssi_value}
                            self.station_test_results[serial_number]["tx_rssi_value"] = {"status": rf_test_status, "value": tx_rssi_value}
                            self.station_test_results[serial_number]["rf_test_status"] = {"status": rf_test_status, "value": ""}
                            # self.root.after(0, partial(self._update_task_ui, station_num, "SIM1 Cellular Test", sim1_cellular_test_status))
                            # time.sleep(0.5)

                            # self.root.after(0, partial(self._update_task_ui, station_num, "SIM2 Cellular Test", sim2_cellular_test_status))
                            # time.sleep(0.5)

                            self.root.after(0, partial(self._update_task_ui, station_num, "GPIO Test", gpio_test_status))
                            time.sleep(0.5)

                            self.root.after(0, partial(self._update_task_ui, station_num, "External Flash Test", external_flash_test_status))
                            time.sleep(0.5)

                            self.root.after(0, partial(self._update_task_ui, station_num, "RF RSSI Test", rf_test_status))
                            time.sleep(0.5)

        return serial_number, test_status, station_num, task_name

    def _execute_shared_test(self, test_name, script_name, result_key, value_key=None):
        """Run one test globally (no station context), apply result to all active stations."""
        self.status_label.config(text=f"Running {test_name} (shared test)...")

        start_time = datetime.now()


        # Collect active stations
        active_stations_for_test = []
        for _, _, station_num in self.active_ports_serials:
            # print(station_num,'station num')
            if station_num not in self.failed_stations and not self.station_stop_flags.get(station_num, False):

                active_stations_for_test.append(station_num)
                # Mark as running
                self.task_indicators[station_num][test_name].config(text="⏳", fg="#2196F3")
                ModernUI.update_status_indicator(self.status_indicators[station_num], "running")
            
            else:
                # Skip failed stations
                self.task_indicators[station_num][test_name].config(text="⛔", fg="#9E9E9E")
                with self.test_completion_lock:
                    self.station_completed_tests.setdefault(test_name, {})[station_num] = True

        self.root.update()
        end_time = datetime.now()

        # If no active stations -> nothing to do
        if not active_stations_for_test:
            with self.test_completion_lock:
                self.completed_tests[test_name] = True
            self.root.after(0, lambda: self.status_label.config(text=f"No active stations for {test_name}"))

            with self.test_completion_lock:
                for com_port, serial_number, station_num in self.active_ports_serials:
                    if serial_number.strip():
                        if test_name == "IMG SINK Programming":
                            # For programming, we already logged inside FirmwareFlash
                            continue
                        if serial_number not in self.station_test_results:
                            self.station_test_results[serial_number] = {}
                        if serial_number not in self.station_test_times:
                            self.station_test_times[serial_number] = {'start': start_time, 'end': end_time}
                        else:
                            self.station_test_times[serial_number]['end'] = end_time
                        self.station_test_results[serial_number][test_name] = {"status": "FAIL", "value": "FAIL"}
                        if test_name == "GPIO Test":       
                            # self.station_test_results[serial_number]["sim1_cellular_test_status"] = {"status": "FAIL", "value": ""}
                            # self.station_test_results[serial_number]["sim2_cellular_test_status"] = {"status": "FAIL", "value": ""}
                            self.station_test_results[serial_number]["gpio_test_status"] = {"status": "FAIL", "value": ""}
                            self.station_test_results[serial_number]["external_flash_test_status"] = {"status": "FAIL", "value": ""}
                            self.station_test_results[serial_number]["rx_rssi_value"] = {"status": "FAIL", "value": ""}
                            self.station_test_results[serial_number]["tx_rssi_value"] = {"status": "FAIL", "value": ""}
                            self.station_test_results[serial_number]["rf_test_status"] = {"status": "FAIL", "value": ""}
                            # self.root.after(0, partial(self._update_task_ui, station_num, "SIM1 Cellular Test", "FAIL"))
                            # self.root.after(0, partial(self._update_task_ui, station_num, "SIM2 Cellular Test", "FAIL"))
                            self.root.after(0, partial(self._update_task_ui, station_num, "GPIO Test", "FAIL"))
                            self.root.after(0, partial(self._update_task_ui, station_num, "External Flash Test", "FAIL"))
                            self.root.after(0, partial(self._update_task_ui, station_num, "RF RSSI Test", "FAIL"))

            return

        try:
            # ✅ Run test ONCE, without station context
            result = self._run_test_script_parallel(
                script_name,
                task_name=test_name,
                result_key=result_key,
                value_key=value_key,
                com_port=None,
                serial_number=None,
                station_num=None
            )
            

            # Default status is FAIL unless explicitly PASS/OK
            status = "FAIL"
            if result and len(result) >= 2:
                _, status, _, _ = result  # Ignore station-specific parts

            # ✅ Apply the SAME result to all stations
            for station_num in active_stations_for_test:
                # self.root.after(0, lambda s=station_num, stat=status: self._update_task_ui(s, test_name, stat))

                if test_name == "Buck Output":
                    pass
                elif test_name == "IMG SINK Programming":
                    pass
                elif test_name == "SUPERCAP DISCHARGED VOLTAGE":
                    pass
                elif test_name == "Super Cap Output":
                    pass
                else:
                    self.root.after(0, lambda s=station_num, stat=status: self._update_task_ui(s, test_name, stat))
                
                if test_name == "IMG SINK Programming":
                    # For programming, we already logged inside FirmwareFlash
                    pass
                else:
                    if status in ["PASS", "OK", "SUCCESS"]:
                        ModernUI.update_status_indicator(self.status_indicators[station_num], "running")
                    else:
                        # self.failed_stations.add(station_num)
                        self.station_stop_flags[station_num] = True
                        ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")

                with self.test_completion_lock:
                    self.station_completed_tests.setdefault(test_name, {})[station_num] = True

        except Exception as e:
            logging.error(f"Shared test {test_name} failed with exception: {e}")
            for station_num in active_stations_for_test:
                self.root.after(0, lambda s=station_num: self._update_task_ui(s, test_name, "FAIL"))
                ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                # self.failed_stations.add(station_num)
                self.station_stop_flags[station_num] = True
                with self.test_completion_lock:
                    self.station_completed_tests.setdefault(test_name, {})[station_num] = True

        # ✅ Mark global test as completed
        with self.test_completion_lock:
            self.completed_tests[test_name] = True
        self.root.after(0, lambda: self.status_label.config(text=f"{test_name} completed (shared)"))

    def _execute_tests_in_parallel(self, test_name, script_name, result_key, value_key=None):
        """Enhanced parallel test execution with ProcessManager and complete station isolation"""

        self.status_label.config(text=f"Running {test_name} tests...")
        
        # Initialize per-station completion tracking for this test
        with self.test_completion_lock:
            if test_name not in self.station_completed_tests:
                self.station_completed_tests[test_name] = {}
            
            # Mark all active stations as not completed for this test
            for _, _, station_num in self.active_ports_serials:
                if station_num not in self.failed_stations:
                    self.station_completed_tests[test_name][station_num] = False


        # Update UI to show tests are running for active stations only
        active_stations_for_test = []
        for optical_com_port, serial_number, station_num in self.active_ports_serials:
            if station_num not in self.failed_stations and not self.station_stop_flags.get(station_num, False):
                active_stations_for_test.append((optical_com_port, serial_number, station_num))
                if test_name == "GPIO Test":
                    self.task_indicators[station_num][test_name].config(text="⏳", fg="#2196F3")
                    # self.task_indicators[station_num]["SIM2 Cellular Test"].config(text="⏳", fg="#2196F3")
                    # self.task_indicators[station_num]["GPIO Test"].config(text="⏳", fg="#2196F3")
                    self.task_indicators[station_num]["External Flash Test"].config(text="⏳", fg="#2196F3")
                    self.task_indicators[station_num]["RF RSSI Test"].config(text="⏳", fg="#2196F3")
                else:    
                    self.task_indicators[station_num][test_name].config(text="⏳", fg="#2196F3")
                ModernUI.update_status_indicator(self.status_indicators[station_num], "running")
            else:
                # Mark failed stations' tests as skipped
                
                if test_name == "GPIO Test":
                    self.task_indicators[station_num][test_name].config(text="⛔", fg="#9E9E9E")
                    # self.task_indicators[station_num]["SIM2 Cellular Test"].config(text="⛔", fg="#9E9E9E")
                    # self.task_indicators[station_num]["GPIO Test"].config(text="⛔", fg="#9E9E9E")
                    self.task_indicators[station_num]["External Flash Test"].config(text="⛔", fg="#9E9E9E")
                    self.task_indicators[station_num]["RF RSSI Test"].config(text="⛔", fg="#9E9E9E")
                else:    
                    self.task_indicators[station_num][test_name].config(text="⛔", fg="#9E9E9E")
                # Mark as completed for failed stations
                with self.test_completion_lock:
                    if test_name in self.station_completed_tests:
                        self.station_completed_tests[test_name][station_num] = True
    
        self.root.update()

        def run_and_update_with_retry(com_port, serial_number, station_num, test_name, max_retries=0):
            """Run test with retry logic and proper error handling - completely isolated per station"""

            
            for attempt in range(max_retries + 1):
                # Skip if station is marked as failed during retry attempts
                if station_num in self.failed_stations or self.station_stop_flags.get(station_num, False):
                    logging.info(f"Station {station_num} marked as failed during retry, skipping {test_name}")
                    # Mark this station's test as completed
                    with self.test_completion_lock:
                        if test_name in self.station_completed_tests:
                            self.station_completed_tests[test_name][station_num] = True
                    return serial_number, "FAIL", station_num, test_name
                
                is_retry = attempt > 0
                if is_retry:
                    logging.info(f"Station {station_num}: Retrying {test_name} (attempt {attempt + 1}/{max_retries + 1})")
                    # Update UI to show retry status
                    self.root.after(0, lambda s=station_num, tn=test_name: self.task_indicators[s][tn].config(text="🔄", fg="#FF9800"))
                    self.root.after(0, lambda s=station_num: ModernUI.update_status_indicator(self.status_indicators[s], "retrying"))
                    time.sleep(3)  # Longer delay before retry for stability
                
                try:
                    # Use the enhanced test script runner
                    # result = self._run_test_script_parallel(script_name, com_port, serial_number, station_num, test_name, result_key, value_key)
                    result = self._run_test_script_parallel(
                        script_name,
                        task_name=test_name,
                        result_key=result_key,
                        value_key=value_key,
                        com_port=com_port,
                        serial_number=serial_number,
                        station_num=station_num
                        )
                
                    if result and len(result) >= 2:
                        sn, status, st_num, task = result
                        
                        if status in ["PASS", "OK"]:
                            # Test passed, update UI and return
                            if test_name == "GPIO Test":
                                pass
                            else:
                                self.root.after(0, lambda s=st_num, t=task, stat=status: self._update_task_ui(s, t, stat))
                            if is_retry:
                                logging.info(f"Station {station_num}: {test_name} succeeded on retry attempt {attempt + 1}")
                            
                            # Mark this station's test as completed
                            with self.test_completion_lock:
                                if test_name in self.station_completed_tests:
                                    self.station_completed_tests[test_name][station_num] = True
                            return result
                        elif status == "FAIL":
                            if attempt < max_retries and test_name != "Memory Test":  # Don't retry Memory Test due to timeout sensitivity
                                # Test failed but we have retries left
                                logging.warning(f"Station {station_num}: {test_name} failed on attempt {attempt + 1}, will retry")
                                continue
                            else:
                                # All retries exhausted or Memory Test failed
                                logging.error(f"Station {station_num}: {test_name} failed after all retry attempts")
                                if test_name == "GPIO Test":
                                    pass
                                else:
                                    self.root.after(0, lambda s=st_num, t=task, stat=status: self._update_task_ui(s, t, stat))
                                
                                # For critical tests, mark station as failed
                                if test_name in ["Test Firmware Download", "NIC Status Test", "Memory Test", "Live Version Test"]:
                                    self.failed_stations.add(station_num)
                                    self.station_stop_flags[station_num] = True
                                
                                # Mark this station's test as completed
                                with self.test_completion_lock:
                                    if test_name in self.station_completed_tests:
                                        self.station_completed_tests[test_name][station_num] = True
                                return result
                            
                except Exception as e:
                    error_msg = f"Exception in {test_name} attempt {attempt + 1}: {str(e)}"
                    logging.error(f"Station {station_num}: {error_msg}")
                    
                    # Kill any hanging processes for this station
                    self.process_manager.kill_station_processes(station_num)
                    
                    if attempt < max_retries and test_name != "Memory Test":
                        continue
                    else:
                        # All retries exhausted due to exceptions
                        self.failed_stations.add(station_num)
                        self.station_stop_flags[station_num] = True
                        
                        # Mark this station's test as completed
                        with self.test_completion_lock:
                            if test_name in self.station_completed_tests:
                                self.station_completed_tests[test_name][station_num] = True
                        return serial_number, "FAIL", station_num, test_name
        
            # Should not reach here, but just in case
            with self.test_completion_lock:
                if test_name in self.station_completed_tests:
                    self.station_completed_tests[test_name][station_num] = True
            return serial_number, "FAIL", station_num, test_name

        # If no active stations, mark test as completed and return
        if not active_stations_for_test:
            with self.test_completion_lock:
                self.completed_tests[test_name] = True
            self.root.after(0, lambda: self.status_label.config(text=f"No active stations for {test_name}"))
            return

        # Execute tests in parallel for active stations only - each station completely independent
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(active_stations_for_test)) as executor:
            future_to_test = {
                executor.submit(run_and_update_with_retry, com_port, serial_number, station_num, test_name): (serial_number, station_num)
                for com_port, serial_number, station_num in active_stations_for_test
                if serial_number.strip()
            }
            
            for future in concurrent.futures.as_completed(future_to_test):
                try:
                    result = future.result()
                    if result:
                        _, status, station_num, _ = result
                        # Update station status indicator based on result
                        if status in ["PASS", "OK"]:
                            self.root.after(0, lambda s=station_num: ModernUI.update_status_indicator(self.status_indicators[s], "running"))
                        elif status == "FAIL":
                            # Only update to failure if this station isn't continuing with other tests
                            if station_num in self.failed_stations:
                                self.root.after(0, lambda s=station_num: ModernUI.update_status_indicator(self.status_indicators[s], "failure"))
                        
                except Exception as e:
                    serial_number, station_num = future_to_test[future]
                    logging.error(f"Exception in {test_name} execution for Station {station_num}: {e}")
                    self.failed_stations.add(station_num)
                    self.station_stop_flags[station_num] = True
                    
                    # Kill any hanging processes for this station
                    self.process_manager.kill_station_processes(station_num)
                    
                    # Mark this station's test as completed
                    with self.test_completion_lock:
                        if test_name in self.station_completed_tests:
                            self.station_completed_tests[test_name][station_num] = True
    
        # Mark the overall test as completed
        with self.test_completion_lock: 
            self.completed_tests[test_name] = True
        self.root.after(0, lambda: self.status_label.config(text=f"{test_name} tests completed"))

    def _update_task_ui(self, station_num, task_name, status):
        """Enhanced UI update with real-time feedback"""
        # print(self.task_indicators,'--------------------------------------------------------------------',task_name)
        if station_num in self.failed_stations:
            return
            
        # Update task indicator with animation effect
        icon = self.task_indicators[station_num][task_name]
        
        if status == "PASS" or status == "OK":
            icon.config(text="✓", fg="#4CAF50")
            # Brief highlight effect
            self.root.after(100, lambda: icon.config(bg="#E8F5E8"))
            self.root.after(500, lambda: icon.config(bg=icon.master["bg"]))
        else:
            icon.config(text="✕", fg="#F44336")
            # Brief highlight effect for failures
            self.root.after(100, lambda: icon.config(bg="#FFEBEE"))
            self.root.after(500, lambda: icon.config(bg=icon.master["bg"]))
        
        # Update status label with current progress
        self.status_label.config(text=f"Station {station_num}: {task_name} - {status}")
        self.root.update()

    def show_start_batch_dialog(self):
        # Dialog to confirm starting a new batch
        dialog = tk.Toplevel(self.root)
        dialog.title("Start New Batch")
        dialog.geometry("400x220")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.configure(bg="white")
        
        # Center the dialog
        dialog.update_idletasks()
        width = dialog.winfo_width()
        height = dialog.winfo_height()
        x = (self.root.winfo_width() // 2) - (width // 2) + self.root.winfo_x()
        y = (self.root.winfo_height() // 2) - (height // 2) + self.root.winfo_y()
        dialog.geometry(f"+{x}+{y}")
        
        # Add a header
        header_frame = tk.Frame(dialog, bg="#2196F3", height=50)
        header_frame.pack(fill=tk.X)
        
        header_label = tk.Label(header_frame, text="Start New Batch", 
                               font=("Arial", 14, "bold"), bg="#2196F3", fg="white")
        header_label.pack(pady=10)
        
        # Dialog content
        content_frame = tk.Frame(dialog, bg="white")
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Create info icon
        icon_canvas = tk.Canvas(content_frame, width=40, height=40, bg="white", highlightthickness=0)
        icon_canvas.pack(pady=5)
        
        # Draw info icon
        icon_canvas.create_oval(5, 5, 35, 35, fill="#E3F2FD", outline="#2196F3", width=2)
        icon_canvas.create_text(20, 20, text="i", font=("Arial", 18, "bold"), fill="#2196F3")
        
        message = tk.Label(content_frame, text="Are you sure you want to start a new batch?", 
                          bg="white", font=("Arial", 11))
        message.pack(pady=10)
        
        # Button frame
        button_frame = tk.Frame(content_frame, bg="white")
        button_frame.pack(pady=10)
        
        # Create buttons
        yes_btn_frame = ModernUI.create_modern_button(button_frame, "Yes", 
                                                    lambda: self.start_new_batch(dialog), 
                                                    width=100, height=30, bg="#4CAF50", hover_bg="#388E3C")
        yes_btn_frame.pack(side=tk.LEFT, padx=10)
        
        no_btn_frame = ModernUI.create_modern_button(button_frame, "No", 
                                                   dialog.destroy, 
                                                   width=100, height=30, bg="#F44336", hover_bg="#D32F2F")
        no_btn_frame.pack(side=tk.LEFT, padx=10)

    def start_new_batch(self, dialog):
        # Close dialog and start a new batch
        dialog.destroy()
        
        # Clean up all processes before restarting
        self.process_manager.cleanup_all_processes()
        
        # Log the reset of counters
        # logging.info(f"Resetting counters. Final values - Total cycles: {self.total_cycles}, Total passed PCBA: {self.total_passed_stations}")
        
        # Destroy the current root window
        self.root.destroy()
        
        # Create a new root window and initialize the GUI again
        new_root = tk.Tk()
        app = ManufacturingSuite(new_root)
        new_root.mainloop()
    
    def find_window_by_title(self, title_keyword):
        """Find window handle by title substring."""
        def enum_windows_callback(hwnd, result):
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                if title_keyword.lower() in title.lower():
                    result.append(hwnd)
        result = []
        win32gui.EnumWindows(enum_windows_callback, result)
        return result[0] if result else None
    
    def launch_minimized(self):
        """Launch the given executable minimized."""
        exe_path = r"C:\Users\Jig Tasting\Desktop\IMG Testing\IMG_V.1.0.0.2\RESEARCHDOWNLOAD_R22.19.1701\Bin\ResearchDownload.exe"
        
        if not os.path.exists(exe_path):
            print("❌ Executable not found:", exe_path)
            return None

        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        si.wShowWindow = win32con.SW_SHOWMINIMIZED

        process = subprocess.Popen([exe_path], startupinfo=si)
        print("✅ Launched minimized:", os.path.basename(exe_path))
        return process
    
    

    def bring_to_foreground(self, window_title="ResearchDownload"):
        """Find the window by title and bring it to the foreground."""
        hwnd = self.find_window_by_title(window_title)
        if hwnd:
            shell = win32com.client.Dispatch("WScript.Shell")
            shell.SendKeys('%')  # Required to allow SetForegroundWindow
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
            print("🟩 Brought to foreground:", window_title)
            return True
        else:
            print("⚠️ Could not find window:", window_title)
            return False

    def stop_test(self):
        """Enhanced stop test with proper cleanup and logging"""
        if self.timer_running:
            if self.timer_id:
                self.root.after_cancel(self.timer_id)
            self.timer_running = False
        
        # Kill all active processes
        self.process_manager.cleanup_all_processes()
        
        # Set all station stop flags
        for station_num in range(1, 3):
            self.station_stop_flags[station_num] = True
        
        # Log test stop event
        for station_num in range(1, 3):
            serial_number = self.station_numbers[station_num].get().strip()
            if serial_number:
                self.excel_logger.log_individual_test(
                    self.session_id, station_num, serial_number, "TEST_STOPPED",
                    datetime.now(), datetime.now(), "User stopped test", 
                     "", "STOPPED", "Test manually stopped by user"
                )
        
        self.status_label.config(text="Test stopped by user.")
        self.status.config(text="OFFLINE")
        self.batch_time.config(text="BATCH TIME: 00:00")

        for station_num in range(1, 3):
            self.station_numbers[station_num].set("")
        
        # Reset status indicators
        for station_num in self.status_indicators:
            ModernUI.update_status_indicator(self.status_indicators[station_num], "pending")

        # Reset task indicators to default state
        for station_num, task_dict in self.task_indicators.items():
            for task, label in task_dict.items():
                label.config(text="⏱", fg="#757575")

        # Clear serial mappings and failed stations
        self.active_ports_serials.clear()
        self.failed_stations.clear()
        
        # Clear test completion tracking
        with self.test_completion_lock:
            for key in self.completed_tests:
                self.completed_tests[key] = False
            self.station_test_results.clear()
            self.station_test_times.clear()
            self.station_completed_tests.clear()

        # Unlock serial entries for re-entry
        for entry in self.entries.values():
            entry.config(state="normal")
            
        # Reset programming phase failed flag
        self.programming_phase_failed = False
        
        # Reset station stop flags
        self.station_stop_flags.clear()

    def _run_all_tests_sequence(self, active_ports_serials):
        """Enhanced test sequence with complete station isolation"""

        logging.info("===== Starting full test sequence =====")
        pythoncom.CoInitialize()
        self.launch_minimized()

        def safe_call(func, *args):
            """Wrapper to ensure one test failure doesn’t stop others"""
            try:
                func(*args)
            except Exception as e:
                logging.error(f"⚠️ Test failed: {args[0]} — {e}", exc_info=True)

        # 1. BOOT Mode Initialization Test
        safe_call(self._execute_shared_test,
                "BOOT Mode Initialization", "staticDO.py", "BOOT_MODE_INITIALIZATION_RESULT", "BOOT_MODE_INITIALIZATION_VALUE")

        # 2. BOOT Mode Test
        safe_call(self._execute_shared_test,
                "BOOT MODE ON", "staticDO.py", "BOOT_Mode_ON_RESULT", "BOOT_Mode_ON_VALUE")

        self.status_label.config(text="Waiting 1 seconds for Supercap charge...")
        self.root.update()
        time.sleep(1)

        # 3. Instant AI Test
        safe_call(self._execute_shared_test,
                "Buck Output", "InstantAI.py", "INSTANT_AI_RESULT", "INSTANT_AI_VALUE")

        # # 3. Static DO2 Off Test
        # safe_call(self._execute_shared_test,
        #         "STATIC DO2 OFF", "staticDO.py", "DO2_OFF_RESULT", "DO2_OFF_VALUE")

        # 4. Static Firmware Flush Test
        safe_call(self._execute_shared_test,
                "IMG SINK Programming", "FirmwareFlash.py", "IMG_SINK_Programming_RESULT", "IMG_SINK_Programming_VALUE")
        




        if self.is_break:
                    # 5. Validate firmware version
            # 5. N58 Programming Initialization
            safe_call(self._execute_shared_test,
                    "N58 Programming Initialization", "", "N58_Programming_Initialization_RESULT", "N58_Programming_Initialization_VALUE")

            
            self.toggle.enable()
            toggle_flag = False

            safe_call(self._execute_shared_test,
                "Super Cap Output", "InstantAI.py", "INSTANT_AI_RESULT", "INSTANT_AI_VALUE")

            safe_call(self._execute_shared_test,
                    "BOOT MODE OFF", "staticDO.py", "BOOT_MODE_OFF_RESULT", "BOOT_MODE_OFF_VALUE")

            safe_call(self._execute_shared_test,
                    "MODULE SUPPLY ON", "staticDO.py", "MODULE_SUPPLY_ON_RESULT", "MODULE_SUPPLY_ON_VALUE")

            # time.sleep(20)
            safe_call(self._execute_tests_in_parallel,
                    "Firmware Version Check", "FirmwareFlash.py", "FIRMWARE_VERSION_CHECK_RESULT", "FIRMWARE_VERSION_CHECK_VALUE")

            # safe_call(self._execute_tests_in_parallel,
            #         "GPIO Test", "test_script.py", "TEST_SCRIPT_RESULT", "TEST_SCRIPT_VALUE")

            safe_call(self._execute_shared_test,
                    "MODULE SUPPLY OFF", "staticDO.py", "MODULE_SUPPLY_OFF_RESULT", "MODULE_SUPPLY_OFF_VALUE")

            safe_call(self._execute_shared_test,
                    "SUPERCAP DISCHARGE", "staticDO.py",
                    "SUPERCAP_DISCHARGE_RESULT", "SUPERCAP_DISCHARGE_VALUE")


            self.status_label.config(text="Waiting 20 seconds for Supercap discharge...")
            self.root.update()
            time.sleep(20)

            safe_call(self._execute_shared_test,
                    "SUPERCAP DISCHARGED VOLTAGE", "InstantAI.py",
                    "SUPERCAP_DISCHARGE_VOLTAGE_RESULT", "SUPERCAP_DISCHARGE_VOLTAGE_VALUE")

            safe_call(self._execute_shared_test,
                    "Module Supply Disconnected", "staticDO.py",
                    "OVERALL_TEST_RESULT", "OVERALL_TEST_RESULT_VALUE")


            logging.info("===== Test sequence completed =====")
            self.root.after(0, self.update_status_after_tests)

        else:
            # 5. N58 Programming Initialization
            safe_call(self._execute_shared_test,
                    "N58 Programming Initialization", "", "N58_Programming_Initialization_RESULT", "N58_Programming_Initialization_VALUE")

            
            self.toggle.enable()
            toggle_flag = False
            time.sleep(5)
            self.bring_to_foreground()
            

            while not self.toggle.is_on:
                continue

            if self.toggle.is_on:
                toggle_flag = True

            if toggle_flag:
                safe_call(self._execute_shared_test,
                "Super Cap Output", "InstantAI.py", "SUPER_CAP_INSTANT_AI_RESULT", "SUPER_CAP_INSTANT_AI_VALUE")
                
                # 5. Validate firmware version
                # safe_call(self._execute_tests_in_parallel,
                #         "Firmware Version Check", "FirmwareFlash.py", "FIRMWARE_VERSION_CHECK_RESULT", "FIRMWARE_VERSION_CHECK_VALUE")
                safe_call(self._execute_shared_test,
                        "BOOT MODE OFF", "staticDO.py", "ALL_DO_OFF_RESULT", "ALL_DO_OFF_VALUE")

                # safe_call(self._execute_shared_test,
                #         "MODULE SUPPLY ON", "staticDO.py", "MODULE_SUPPLY_ON_RESULT", "MODULE_SUPPLY_ON_VALUE")
                # time.sleep(5)
                safe_call(self._execute_shared_test,
                    "MODULE SUPPLY ON", "staticDO.py", "MODULE_SUPPLY_ON_RESULT", "MODULE_SUPPLY_ON_VALUE")
                # time.sleep(20)
                safe_call(self._execute_tests_in_parallel,
                        "Firmware Version Check", "FirmwareFlash.py", "FIRMWARE_VERSION_CHECK_RESULT", "FIRMWARE_VERSION_CHECK_VALUE")
                # time.sleep(5)

                # safe_call(self._execute_tests_in_parallel,
                #         "GPIO Test", "test_script.py", "TEST_SCRIPT_RESULT", "TEST_SCRIPT_VALUE")

                safe_call(self._execute_shared_test,
                        "MODULE SUPPLY OFF", "staticDO.py", "MODULE_SUPPLY_OFF_RESULT", "MODULE_SUPPLY_OFF_VALUE")

                # safe_call(self._execute_tests_in_parallel,
                #         "GPIO Test", "test_script.py", "TEST_SCRIPT_RESULT", "TEST_SCRIPT_VALUE")

                # safe_call(self._execute_shared_test,
                #         "MODULE SUPPLY OFF", "staticDO.py", "MODULE_SUPPLY_OFF_RESULT", "MODULE_SUPPLY_OFF_VALUE")

                safe_call(self._execute_shared_test,
                        "SUPERCAP DISCHARGE", "staticDO.py",
                        "SUPERCAP_DISCHARGE_RESULT", "SUPERCAP_DISCHARGE_VALUE")


                self.status_label.config(text="Waiting 20 seconds for Supercap discharge...")
                self.root.update()
                time.sleep(20)

                safe_call(self._execute_shared_test,
                        "SUPERCAP DISCHARGED VOLTAGE", "InstantAI.py",
                        "SUPERCAP_DISCHARGE_VOLTAGE_RESULT", "SUPERCAP_DISCHARGE_VOLTAGE_VALUE")

                safe_call(self._execute_shared_test,
                        "Module Supply Disconnected", "staticDO.py",
                        "OVERALL_TEST_RESULT", "OVERALL_TEST_RESULT_VALUE")


                logging.info("===== Test sequence completed =====")
                self.root.after(0, self.update_status_after_tests)

    def update_timer(self):
        """Update the batch time display"""
        if self.start_time and self.timer_running:
            elapsed_time = time.time() - self.start_time
            minutes = int(elapsed_time // 60)
            seconds = int(elapsed_time % 60)
            self.batch_time.config(text=f"BATCH TIME: {minutes:02d}:{seconds:02d}")
            self.timer_id = self.root.after(1000, self.update_timer)

    def run_test(self):
        """Enhanced run test with comprehensive logging, progress tracking, and complete station isolation"""

        # Reset UI and internal states
        for station_num in range(1, 5):
            self.station_numbers[station_num].set("")
        
        for station_num in self.status_indicators:
            ModernUI.update_status_indicator(self.status_indicators[station_num], "pending")

        for station_num, task_dict in self.task_indicators.items():
            for task, label in task_dict.items():
                label.config(text="⏱", fg="#757575")

        self.active_ports_serials.clear()
        self.failed_stations.clear()
        
        with self.test_completion_lock:
            for key in self.completed_tests:
                self.completed_tests[key] = False
            self.station_test_results.clear()
            self.station_test_times.clear()
            self.station_completed_tests.clear()

        for entry in self.entries.values():
            entry.config(state="normal")

        # Reset station stop flags and programming phase flag
        self.station_stop_flags.clear()
        self.programming_phase_failed = False

        self.status.config(text="ONLINE")

        # Start timer
        self.start_time = time.time()
        self.timer_running = True
        self.update_timer()
        
        self.status_label.config(text="Test is running...")

        # Log test session start
        for station_num in range(1, 5):
            self.excel_logger.log_test_start(self.session_id, station_num, "PENDING_SCAN")

        # Step 0: Auto-scan serial numbers
        self.status_label.config(text="Starting auto-scan for serial numbers...")
        self.auto_scan_all_serials()
        self.root.update()
        time.sleep(2)

        # Count active stations and update status indicators
        valid_stations_to_test = []
        for station_num in range(1, 5):
            if station_num not in self.failed_stations:
                serial_number = self.station_numbers[station_num].get().strip()
                if serial_number:
                    valid_stations_to_test.append(station_num)
                    ModernUI.update_status_indicator(self.status_indicators[station_num], "running")
        
        if not valid_stations_to_test:
            self.status_label.config(text="No valid stations to test - all stations failed serial scan or config.")
            logging.warning("No valid stations to test - all stations failed serial scan or config.")
            self.stop_test()
            return
        
        self.root.update()

        # Step 1: Map Serial Number with COM Ports
        self.gather_com_ports_serial_numbers()
        self.root.update()
        time.sleep(0.5)


        # Filter active_ports_serials to only include valid stations
        self.active_ports_serials = [
            (com, sn, st_num) for com, sn, st_num in self.active_ports_serials
            if st_num not in self.failed_stations
        ]
        
        if not self.active_ports_serials:
            self.status_label.config(text="No stations with valid COM port mapping to proceed with tests.")
            logging.warning("No stations with valid COM port mapping to proceed with tests.")
            self.stop_test()
            return
        
        # Now run the sequence of subsequent tests in a separate thread

        
        threading.Thread(target=self._run_all_tests_sequence,args=(self.active_ports_serials,), daemon=True).start()
        
        logging.info(f"Test sequence initiated for session {self.session_id}. Final status will update upon completion of all tests.")

    def check_all_pass(self, data):
        if not data:
            return False
        if all((test.get("status") == "PASS" )or (test.get("status") == "OK") for test in data.values()):
            return True
        else:
            return False

    def update_status_after_tests(self):
        """Enhanced status update with proper handling of failed stations and complete isolation"""
        logging.debug(f"Entering update_status_after_tests. Current completed_tests state: {self.completed_tests}")

        with self.test_completion_lock:
            # Check if all tests are completed
            
            self.completed_tests['Mapped with COM Port'] = True
            all_tests_completed = all(self.completed_tests.values())
            
            if not all_tests_completed:
                logging.debug("Not all tests completed yet. Rescheduling update_status_after_tests.")
                self.root.after(200, self.update_status_after_tests)
                return

        logging.info("All tests completed. Finalizing...")
        overall_pass_count = 0
        overall_fail_count = 0

        for station_num in range(1, 5):  # Assuming 4 stations
            serial_number = self.station_numbers[station_num].get().strip()
            if self.eng_var.get() and " - " in serial_number:
                serial_number = serial_number.split(" - ")[-1]

            if not serial_number or station_num in self.failed_stations:
                if station_num in self.failed_stations:
                    overall_fail_count += 1
                    ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")

            station_results = self.station_test_results.get(serial_number, {})
            critical_tests_passed = 0
            total_tests_run = 0
            all_tests_passed = True  # Flag to track if all tests passed

            # Check all test results for this station

            for test_name, test_result in station_results.items():
                if test_name == 'IMG SINK Programming':
                    continue
                total_tests_run += 1
                if test_result.get('status', 'FAIL') in ['PASS', 'OK']:
                    critical_tests_passed += 1
                else:
                    all_tests_passed = False  # If any test fails, mark as not all passed

            pass_percentage = (critical_tests_passed / total_tests_run) if total_tests_run > 0 else 0
            firmware_passed = station_results.get('IMG SINK Programming', {}).get('status', 'FAIL') in ['OK', 'PASS']

            # Determine overall station status
            if station_num in self.station_stop_flags:
                overall_station_status = "FAIL"
                ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                overall_fail_count += 1
            else:
                overall_station_status = self.check_all_pass(self.station_test_results.get(serial_number))
                if overall_station_status:
                    overall_station_status = "PASS"
                    ModernUI.update_status_indicator(self.status_indicators[station_num], "success")
                    overall_pass_count += 1
                
                else:
                    overall_station_status = "FAIL"
                    ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
                    overall_fail_count += 1
            # elif firmware_passed and pass_percentage >= 0.7 and all_tests_passed:
            #     overall_station_status = "PASS"
            #     ModernUI.update_status_indicator(self.status_indicators[station_num], "success")  # Set to success
            #     overall_pass_count += 1
            # else:
            #     overall_station_status = "FAIL"
            #     ModernUI.update_status_indicator(self.status_indicators[station_num], "failure")
            #     overall_fail_count += 1

            station_results['overall_result'] = {'status': overall_station_status}
            logging.info(f"Station {station_num} ({serial_number}) final result: {overall_station_status}")
            self.excel_logger.log_final_results(
                self.session_id, station_num, serial_number, station_results,
                self.station_test_times.get(serial_number, {}).get('start', datetime.now()),
                self.station_test_times.get(serial_number, {}).get('end', datetime.now())
            )


        self.total_cycles += 1
        self.total_passed_stations += overall_pass_count
        logging.debug(f"Updated total_cycles = {self.total_cycles}")
        logging.debug(f"Updated total_passed_stations = {self.total_passed_stations}")

        # Update cycle and pass labels
        try:
            # if self.cycle_count_label:
            #     self.cycle_count_label.config(text=f"Total Cycles: {self.total_cycles}")
            #     self.cycle_count_label.update_idletasks()
            # else:
            #     logging.warning("cycle_count_label is None")

            # if self.passed_stations_label:
            #     self.passed_stations_label.config(text=f"Passed PCBA: {self.total_passed_stations}")
            #     self.passed_stations_label.update_idletasks()
            # else:
            #     logging.warning("passed_stations_label is None")

            if self.status:
                self.status.config(text="OFFLINE")
                self.status.update_idletasks()
            overall_fail_count = len(self.station_numbers) - overall_pass_count
            summary_msg = f"Test completed: {overall_pass_count} PASS, {overall_fail_count} FAIL"
            skipped_count = len(self.station_numbers) - (overall_pass_count + overall_fail_count)
            if skipped_count > 0:
                summary_msg += f", {skipped_count} SKIPPED."

            if self.status_label:
                self.status_label.config(text=summary_msg)
                self.status_label.update_idletasks()
            else:
                logging.warning("status_label is None")
        except Exception as e:
            logging.error(f"Label update failed: {e}")

        if self.timer_id:
            self.root.after_cancel(self.timer_id)
            self.timer_id = None
            logging.info("Batch timer stopped after test completion.")

        self.timer_running = False

if __name__ == "__main__":
    root = tk.Tk()
    app = ManufacturingSuite(root)
    root.mainloop()
