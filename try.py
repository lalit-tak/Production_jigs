
import tkinter as tk
from tkinter import ttk, messagebox
import serial
import serial.tools.list_ports
import re
import subprocess
import time
import concurrent.futures
import threading
import json
import logging
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk, ImageDraw, ImageFont
from firmware_version_integration import check_firmware_version
import tkinter.simpledialog as simpledialog
from zipfile import BadZipFile
import ast

# Get current date for folder naming and filename
current_date = datetime.now().strftime("%Y-%m-%d")

# Setup directories
Test_Result = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Test Result Excel File")
Log_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Log files")

os.makedirs(Test_Result, exist_ok=True)
os.makedirs(Log_folder, exist_ok=True)

log_folder = os.path.join(Log_folder, f"Log_file_{current_date}")
excel_folder = os.path.join(Test_Result, f"Test_Result_{current_date}")

os.makedirs(log_folder, exist_ok=True)
os.makedirs(excel_folder, exist_ok=True)

# file_name = os.path.join(excel_folder, f"{current_date}_NIC_RF_TEST_RESULT.xlsx")

# if os.path.exists(file_name):
#     workbook = load_workbook(file_name)
#     sheet = workbook.active
# else:
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.append(["NIC serial number", "RF test result", "RX RSSI", "TX RSSI", "Action", "Timestamp"])

# Build Excel file name
file_name = os.path.join(excel_folder, f"{current_date}_NIC_RF_TEST_RESULT.xlsx")

# Try loading workbook if it exists, else create new one
try:
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        raise FileNotFoundError  # Trigger new workbook creation
except (BadZipFile, FileNotFoundError):
    print(f"⚠️ Creating new Excel file. Reason: {'Corrupted file' if os.path.exists(file_name) else 'File not found'}")
    if os.path.exists(file_name):
        # Optionally rename the corrupted file
        os.rename(file_name, file_name + ".corrupted")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIC serial number", "Module Supply On", "Supercap Charging Voltage", "Supercap Charging Voltage Status", "LDO Voltage", "LDO Voltage Status", "RF test result", "RX RSSI", "TX RSSI", "RF Status", "External Flash", "Firmware Version", "Expected Firmware Version", "Firmware Version Test Status", "Module Supply Off", "Supercap Discharging Voltage", "Supercap Discharging Voltage Status", "Overall Result"])

def insert_into_excel(data_dict):
    # print(data_dict)
    row_data = []
    for station_num in list(data_dict.keys()):
        module_supply_on = data_dict[(station_num)].get('Module_Supply_On', {})
        Super_Cap_Charge_Voltage = data_dict[(station_num)].get('Super Cap Charge Voltage', {})
        LDO_Voltage = data_dict[(station_num)].get('LDO Voltage', {})
        Module_Supply_Off = data_dict[(station_num)].get('Module_Supply_Off', {})
        Super_Cap_Discharge_Voltage = data_dict[(station_num)].get('Super Cap Discharge Voltage', {})

        row_data = [
            data_dict[(station_num)].get('serial_number', ''),
            module_supply_on.get('status', "FAIL"),
            Super_Cap_Charge_Voltage.get('value'),
            Super_Cap_Charge_Voltage.get('status'),
            LDO_Voltage.get('value'),
            LDO_Voltage.get('status'),
            data_dict[(station_num)].get('rf_test_result', 'FAIL'),
            data_dict[(station_num)].get('tx_rssi', 'N/A'),
            data_dict[(station_num)].get('rx_rssi', 'N/A'),
            data_dict[(station_num)].get('rf_status', 'FAIL'),
            data_dict[(station_num)].get('flash_status', 'FAIL'),
            data_dict[(station_num)].get('Firmware_Version', ''),
            data_dict[(station_num)].get('Expected_Firmware_Version', ''),
            data_dict[(station_num)].get('Check_Firmware_Version', 'FAIL'),
            Module_Supply_Off.get('status','FAIL'),
            Super_Cap_Discharge_Voltage.get('value', ''),
            Super_Cap_Discharge_Voltage.get("status", 'FAIL'),
            data_dict[(station_num)].get('Overall_result', 'FAIL')
        ]
        

        sheet.append(row_data)
        workbook.save(file_name)
    
    workbook.save(file_name)

data_dict = {4: {'serial_number': 'P25100301000006900300371972', 'Module_Supply_On': {'status': 'PASS'}, 'Super Cap Charge Voltage': {'status': 'FAIL', 'value': '1.395263671875'}, 'LDO Voltage': {'status': 'FAIL', 'value': '1.402587890625'}, 'Module_Supply_Off': {'status': 'PASS'}, 'Check_Firmware_Version': False, 'Firmware_Version': 'None', 'Expected_Firmware_Version': '6.1.0.0', 'Super Cap Discharge Voltage': {'status': 'FAIL', 'value': '1.397705078125'}},
1: {'serial_number': 'P25100301000006900300380020', 'Module_Supply_On': {'status': 'PASS'}, 'Super Cap Charge Voltage': {'status': 'PASS', 'value': '3.187255859375'}, 'LDO Voltage': {'status': 'PASS', 'value': '3.310546875'}, 'Module_Supply_Off': {'status': 'PASS'}, 'Check_Firmware_Version': False, 'Firmware_Version': 'None', 'Expected_Firmware_Version': '6.1.0.0', 'Super Cap Discharge Voltage': {'status': 'PASS', 'value': '3.726806640625'}},
5: {'serial_number': 'P25100301000006900300368907', 'Module_Supply_On': {'status': 'PASS'}, 'Super Cap Charge Voltage': {'status': 'PASS', 'value': '3.33984375'}, 'LDO Voltage': {'status': 'PASS', 'value': '3.31298828125'}, 'Module_Supply_Off': {'status': 'PASS'}, 'Check_Firmware_Version': False, 'Firmware_Version': 'None', 'Expected_Firmware_Version': '6.1.0.0', 'Super Cap Discharge Voltage': {'status': 'PASS', 'value': '3.8525390625'}},
6: {'serial_number': 'P25100301000006900300368754', 'Module_Supply_On': {'status': 'PASS'}, 'Super Cap Charge Voltage': {'status': 'PASS', 'value': '3.187255859375'}, 'LDO Voltage': {'status': 'PASS', 'value': '3.310546875'}, 'rf_test_result': 'Fail', 'tx_rssi': 'N/A', 'rx_rssi': 'N/A', 'rf_status': 'Fail', 'flash_status': 'Fail', 'Module_Supply_Off': {'status': 'PASS'}, 'Check_Firmware_Version': False, 'Firmware_Version': 'None', 'Expected_Firmware_Version': '6.1.0.0', 'Super Cap Discharge Voltage': {'status': 'PASS', 'value': '3.731689453125'}, 'Overall_result': None},
2: {'serial_number': 'P25100301000006900300368906', 'Module_Supply_On': {'status': 'PASS'}, 'Super Cap Charge Voltage': {'status': 'PASS', 'value': '3.189697265625'}, 'LDO Voltage': {'status': 'PASS', 'value': '3.314208984375'}, 'Module_Supply_Off': {'status': 'PASS'}, 'Check_Firmware_Version': False, 'Firmware_Version': 'None', 'Expected_Firmware_Version': '6.1.0.0', 'Super Cap Discharge Voltage': {'status': 'PASS', 'value': '3.720703125'}},
3: {'serial_number': 'P25100301000006900300369681', 'Module_Supply_On': {'status': 'PASS'}, 'Super Cap Charge Voltage': {'status': 'FAIL', 'value': '1.396484375'}, 'LDO Voltage': {'status': 'FAIL', 'value': '1.392822265625'}, 'Module_Supply_Off': {'status': 'PASS'}, 'Check_Firmware_Version': False, 'Firmware_Version': 'None', 'Expected_Firmware_Version': '6.1.0.0', 'Super Cap Discharge Voltage': {'status': 'FAIL', 'value': '1.39892578125'}}}


insert_into_excel(data_dict)